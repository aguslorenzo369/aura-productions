# -*- coding: utf-8 -*-
"""Construye el dataset consolidado a partir de los dos xlsx originales."""
import openpyxl, datetime, json, unicodedata

HOY = datetime.date(2026, 9, 2)

def d(s):
    try: return datetime.datetime.strptime(str(s).strip(), '%d/%m/%Y').date()
    except Exception: return None

# ---------- 1. CRONOGRAMA ----------
cr = openpyxl.load_workbook('cronograma.xlsx', data_only=True)
calendario = []
for r in cr['CALENDARIO DE PAGOS'].iter_rows(min_row=4, values_only=True):
    f = d(r[0])
    if not f or r[1] is None: continue
    calendario.append(dict(fecha=f, pais=str(r[1]).strip(), cat=str(r[2]).strip(),
                           concepto=str(r[3]).strip(), cuota=str(r[4]).strip(),
                           monto=float(r[5]), estado=str(r[6]).strip()))
historico = []
for r in cr['PAGOS REALIZADOS'].iter_rows(min_row=4, values_only=True):
    f = d(r[0])
    if not f or r[5] is None: continue
    historico.append(dict(fecha=f, pais=str(r[1]).strip(), cat=str(r[2]).strip(),
                          concepto=str(r[3]).strip(), pct=str(r[4]), monto=float(r[5])))

# ---------- 2. MASTER: hoja PAGOS (contratos / reservas) ----------
inv = openpyxl.load_workbook('inversion.xlsx', data_only=True)
ws = inv['PAGOS']
contratos = []
for i, r in enumerate(ws.iter_rows(min_row=5, max_row=60, values_only=True), 5):
    pais, cat, prov, tot = r[2], r[4], r[5], r[6]
    if not prov: continue
    num = lambda v: float(v) if isinstance(v, (int, float)) else None
    contratos.append(dict(
        fila=i,
        pais=str(pais).strip() if pais else '',
        cat=str(cat).strip() if cat else '',
        proveedor=str(prov).strip(),
        total=num(tot),
        pct1=r[7], pago1=num(r[8]), pago1_txt=r[8] if isinstance(r[8], str) else None,
        saldo=num(r[9]),
        fecha_saldo=r[10].date() if isinstance(r[10], datetime.datetime) else None,
        pct2=r[11], pago2=num(r[12]),
        fecha_reserva=r[1].date() if isinstance(r[1], datetime.datetime) else None,
    ))

# ---------- 3. MASTER: hojas por evento ----------
FX = {  # hoja -> (moneda, factor a USD, nota)
    'Ecuador (25 Jul-26 Jul)':      ('USD', 1.0,        'USD directo'),
    'Panama (09 ago-10 ago)':       ('USD', 1.0,        'USD directo'),
    'Rep D. (15 ago-16 ago)':       ('USD', 1.0,        'Hoja ya en USD (tasa 1.0)'),
    'Colombia (29 ago - 30 ago)':   ('COP', 1/3700,     'Presupuesto /3700; la columna PAGADO usa /4080'),
    'Mexico (5 sep - 6 sep)':       ('MXN', 0.059,      'Presupuesto x0.059; la columna PAGADO usa x0.051'),
    'Chile (12 sep-13 sep)':        ('USD', 1.0,        'Celda de tasa vacia/0 -> se toma como USD'),
    'Uruguay (19 sep-20 sep)':      ('USD', 1.0,        'Celda de tasa 0 -> se toma como USD'),
    'Costa Rica (1 ago-2 ago)':     ('USD', 1.0,        'Sin celda de tasa'),
    'Argentina (3 oct-4 oct)':      ('USD', 1.0,        'Tasa 1.0'),
    'Salvador (7 nov-8 nov)':       ('USD', 1.0,        'Tasa heredada de Mexico (link roto)'),
    'Guatemala (14 nov-15 nov)':    ('USD', 1.0,        'Tasa heredada de Mexico (link roto)'),
    'Perú ':                        ('USD', 1.0,        'Tasa heredada de Mexico (link roto)'),
    'España (28 nov-29 nov)':       ('EUR', 1.15,       'EUR x1.15'),
}
eventos = []
for hoja, (moneda, factor, nota) in FX.items():
    ws = inv[hoja]
    fila_tot = None
    for i, r in enumerate(ws.iter_rows(min_row=1, max_row=120, values_only=True), 1):
        c0 = str(r[0]).strip().upper() if r[0] else ''
        if c0.startswith('TOTAL'): fila_tot = i
    vals = [ws.cell(fila_tot, j).value for j in range(1, ws.max_column + 1)]
    nums = [v for v in vals if isinstance(v, (int, float))]
    # Total = penultimo bloque; usamos localizacion por encabezado
    hdr = {}
    for j in range(1, ws.max_column + 1):
        h = ws.cell(7, j).value
        if h: hdr[str(h).strip()] = j
    col_tot = hdr.get('Total') or hdr.get('Totales')
    col_pago = hdr.get('Pago')
    total_local = ws.cell(fila_tot, col_tot).value if col_tot else None
    pago_local = ws.cell(fila_tot, col_pago).value if col_pago else None
    # cabecera del evento (fecha interna)
    cab = ''
    for i in range(1, 7):
        for j in range(1, 8):
            v = ws.cell(i, j).value
            if v and 'Check List' in str(v): cab = ' '.join(str(v).split())
    eventos.append(dict(hoja=hoja, moneda=moneda, factor=factor, nota_fx=nota,
                        fila_totales=fila_tot,
                        total_local=float(total_local) if isinstance(total_local, (int, float)) else None,
                        pago_local=float(pago_local) if isinstance(pago_local, (int, float)) else None,
                        cabecera=cab))

json.dump(dict(
    calendario=[{**c, 'fecha': str(c['fecha'])} for c in calendario],
    historico=[{**h, 'fecha': str(h['fecha'])} for h in historico],
    contratos=[{**c, 'fecha_saldo': str(c['fecha_saldo']) if c['fecha_saldo'] else '',
                'fecha_reserva': str(c['fecha_reserva']) if c['fecha_reserva'] else ''} for c in contratos],
    eventos=eventos,
), open('consolidado.json', 'w'), ensure_ascii=False, indent=1, default=str)
print('calendario', len(calendario), '| historico', len(historico), '| contratos', len(contratos), '| eventos', len(eventos))
for e in eventos:
    tl, pl = e['total_local'], e['pago_local']
    print('  %-30s %-4s total_local %14s -> USD %12.2f | pagado_local %12s -> USD %12.2f' % (
        e['hoja'][:30], e['moneda'], round(tl,2) if tl else '-', (tl or 0)*e['factor'],
        round(pl,2) if pl else '-', (pl or 0)*e['factor']))
