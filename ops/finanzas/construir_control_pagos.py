# -*- coding: utf-8 -*-
"""Datos del CONTROL DE PAGOS A PROVEEDORES (v2), transcriptos del Google Sheet
12w0kcBaHNuSMdFAJi0PmpdJOSpRlk-5DuwmPXs2lqyI el 03/09/2026."""

# (proveedor, concepto, plaza, estado, moneda, neto, iva, total, sena_pct, sena,
#  saldo, usd, via, razon, cuit, domicilio, titular, banco, cbu, alias, swift, invoice, nota)
N = None
FILAS = [
 ('La Rural S.A.', 'Locación Pabellón Azul — Contrato CC-00957', 'Argentina', 'Contratado', 'ARS',
  N,N,N,N,N,N,N, '', '', '', '', '', '', '', '', '', '',
  'FALTA MONTO. Contrato CC-00957. 10.320 m². 2-3-4 oct 2026. No hay invoice IIDAI ni registro de pago del contrato del predio.'),
 ('La Rural — Infraestructura', 'Infraestructura y mobiliario (Arq. Christian Riccio)', 'Argentina',
  'Presupuesto recibido sin responder', 'ARS', N,N,N,N,N,N,N, '', '', '', '', '', '', '', '', '', '',
  'FALTA MONTO — está en el PDF del mail del 17/08. Guardar en Drive para cargarlo.'),
 ('La Rural — Conectividad', 'WiFi 2 redes privadas (recotización 100 Mbps)', 'Argentina',
  'Esperando recotización', 'ARS', N,N,N,N,N,N,N, '', '', '', '', '', '', '', '', '', '',
  'FALTA MONTO — presupuesto del 16/06 en PDF. Recotización a 100 Mbps pedida el 05/08, sin respuesta.'),
 ('Grupo MET', 'Producción técnica + entelado', 'Argentina', 'Cerrado + señado', 'USD',
  N,N,N,N, 5000.0, N, 5000.0, 'Mercury / transferencia', 'Grupo MET LLC', '', '', 'Grupo MET LLC',
  '', '', '', '', '',
  'FALTA MONTO TOTAL. Pagado USD 5.000 vía Mercury (INV-11, 31/07). Contacto: Lalo. Sus mails salientes fallaban: confirmar por WhatsApp. Descuento ~15M ARS.'),
 ('Grupo Ambient / Good Events SA', 'Catering desayuno VIP 600 pax — Presupuesto S03362', 'Argentina',
  'Cerrado — esperando seña', 'ARS', 12522600.0, 2629746.0, 15152346.0, 0.50, 7576173.0, 7576173.0, N,
  'Transferencia', 'Good Events SA', '', '', 'Good Events SA', '', '', '', '', '2601021',
  'CORREGIDO A 600 pax × $20.871 + IVA. El presupuesto S03362 y la invoice 2601021 se emitieron por 800: Ángeles García Laborde no hizo el ajuste. RECLAMAR la reemisión antes de señar. Cuenta a nombre de Good Events SA, NO Gourmet Connection SA.'),
 ('Grupo Ambient / Good Events SA', 'Catering staff 150 pax — Presupuesto S03363', 'Argentina',
  'Cerrado — esperando seña', 'ARS', 8258100.0, 1734201.0, 9992301.0, 0.50, 4996150.5, 4996150.5, N,
  'Transferencia', 'Good Events SA', '', '', 'Good Events SA', '', '', '', '', '2601021',
  '150 pax × $55.054 + IVA. Anticipo unificado con S03362.'),
 ('Grupo Ambient — TOTAL UNIFICADO', 'Anticipo 50% de S03362 + S03363', 'Argentina',
  'Cerrado — esperando seña', 'ARS', 10390350.0, 2181973.5, 12572323.5, 0.50, 12572323.5, 12572323.5, 8190.44,
  'Transferencia', 'Good Events SA', '', '', 'Good Events SA', '', '', '', '', '2601021',
  'ANTICIPO CORREGIDO a 600 desayunos: base neta total $20.780.700. OJO: la invoice 2601021 emitida dice $15.097.714,50 porque está calculada sobre 800 desayunos. Si se paga como está se transfieren $2.525.391 de más (US$1.645). Pedir la reemisión. Saldo 50% pendiente de emitir.'),
 ('Chanes Seguros', 'Seguro de accidentes personales + asistencia médica 150 pax', 'Argentina',
  'Cerrado — esperando seña', 'ARS', N, N, 953151.93, N, N, 953151.93, N, 'Transferencia',
  '', '', '', '', '', '', '', '', '',
  'MONTO CONFIRMADO por mail del 13/07. Confirmar si incluye IVA. Falta póliza emitida.'),
 ('Chanes Seguros', 'Seguro de RESPONSABILIDAD CIVIL', 'Argentina', 'NUNCA LLEGÓ', 'ARS',
  N,N,N,N,N,N,N, 'Transferencia', '', '', '', '', '', '', '', '', '',
  'Néstor prometió el 13/07 pasar el costo. Nunca llegó y La Rural suele exigir RC. RECLAMAR.'),
 ('Higia Eventos', 'Limpieza sábado y domingo 07:00–23:30', 'Argentina',
  'Cerrado — esperando presupuesto ajustado', 'ARS', N,N,N,N,N,N,N, 'Transferencia',
  '', '', '', '', '', '', '', '', '',
  'FALTA MONTO — presupuesto de Yanina Cruz del 22/06 en PDF. Ajuste pedido el 11/08, sin respuesta. Incluir desayuno VIP del domingo 8am.'),
 ('Gale Servicios', 'Retiro y gestión de residuos', 'Argentina', 'Cerrado — esperando seña', 'ARS',
  N,N,N,N,N,N,N, 'Transferencia', '', '', '', '', '', '', '', '', '',
  'FALTA MONTO — cotización de Yanina Cruz del 22/06 en PDF. Proveedor exclusivo del predio. Mismo grupo que Higia.'),
 ('Road Seguridad', 'Seguridad y vigilancia 2-3-4 oct', 'Argentina', 'Negociación', 'ARS',
  N,N,N,N,N,N,N, 'Transferencia', 'Road Seguridad S.A.', '',
  'Esteban Echeverría 2864 — Munro B1605DSR — Buenos Aires', 'Road Seguridad S.A.', '', '', '', '', '',
  'FALTA MONTO — está en el XLS del 12/08. OJO: el archivo se llama "3 Y 4 DE OCTUBRE" y el esquema incluye el viernes 2 (montaje). VERIFICAR que cotizaron los 3 días. Contacto: Javier Celiz. Tarifa de junio: actualizar a octubre.'),
 ('Vittal (Socorro Médico Privado SA)', 'Servicio médico — ambulancia UTIM + 2 médicos + enfermero',
  'Argentina', 'Negociación — a recotizar', 'ARS', 6920220.32, 726623.13, 7646843.45, N, N, N, N,
  'Transferencia 100% adelantado', 'Socorro Médico Privado S.A.', '',
  'Av. Álvarez Thomas 1154 (C1427CCY) CABA', 'Socorro Médico Privado S.A.', '', '', '', '', '',
  'Referencia: cotización del 18/06 CON consultorio. VENCIDA y ahora SIN consultorio: el número debe BAJAR. Contacto: Vanesa Tomas. Pedir ART + matrículas + renuncia de repetición.'),
 ('Mobiliario / sillas', '5.500 sillas (1.000 tapizadas VIP + 4.500 plásticas) + salas', 'Argentina',
  'Negociación', 'ARS', N,N,N,N,N,N,N, 'A definir', '', '', '', '', '', '', '', '', '',
  'FALTA MONTO — presupuesto FDL del 07/08. FDL NO factura al exterior (usa un tercero que emite en USA). Casablanca sólo las 1.000 tapizadas. Sealquilatodo sin respuesta.'),
 ('Baños químicos', '25 baños químicos exteriores', 'Argentina', 'Bloqueado', '',
  N,N,N,N,N,N,N, '', '', '', '', '', '', '', '', '', '',
  'Pendiente del plano del predio. Sin proveedor asignado.'),
 ('Food trucks', 'Food trucks zona exterior', 'Argentina', 'Bloqueado', '',
  N,N,N,N,N,N,N, '', '', '', '', '', '', '', '', '', '',
  'La Rural nunca indicó con quién gestionar la autorización (pedido el 24/06).'),
 ('Colorvox', 'Regalo VIP CMC 2026 (contenedor con y sin tarjeta)', 'Argentina',
  'Cotización recibida', '', N,N,N,N,N,N,N, '', '', '', '', '', '', '', '', '', '',
  'FALTA MONTO — dos cotizaciones del 24/08 en PDF. Contacto: Alejandra Corzo / Héctor Caldera.'),
 ('REMERASYESTAMPADOS', 'Remeras × $14.110 — PAGO TOTAL 100%', 'Argentina', 'Facturado', 'ARS',
  3527500.0, 740775.0, 4268275.0, 1.00, 4268275.0, 0.0, 2780.64, 'USDT BEP-20',
  'Leandro Daniel Petracci', '', 'Uruguay 41 — Villa Martelli CP 1603', 'Leandro Daniel Petracci',
  'Galicia', '0070056630004018140418', '', 'GABAARBA', '2601017',
  'Versión FINAL en la carpeta "Nuevo pago -". OJO: la planilla decía 230 unidades y la factura 2601017 dice 250; el neto coincide con 250.'),
 ('REMERASYESTAMPADOS', 'Pañuelos — PAGO 50%', 'Argentina', 'Facturado (50%)', 'ARS',
  13387500.0, 2811375.0, 16198875.0, 0.50, 8099437.5, 8099437.5, 5276.51, 'USDT BEP-20',
  'Leandro Daniel Petracci', '', 'Uruguay 41 — Villa Martelli CP 1603', 'Leandro Daniel Petracci',
  'Galicia', '0070056630004018140418', '', 'GABAARBA', '2601019',
  'Saldo 50% de $8.099.437,50 PENDIENTE DE EMITIR. Compra compartida con Uruguay.'),
 ('REMERASYESTAMPADOS', 'Bordado de gorras — PAGO TOTAL 100%', 'Argentina', 'Facturado', 'ARS',
  2643200.0, 0.0, 2643200.0, 1.00, 2643200.0, 0.0, 1721.95, 'USDT BEP-20',
  'Leandro Daniel Petracci', '', 'Uruguay 41 — Villa Martelli CP 1603', 'Leandro Daniel Petracci',
  'Galicia', '0070056630004018140418', '', 'GABAARBA', '2601016',
  'SIN IVA (precio final). Cotización original $3.055.000 → final $2.643.200. Se eligió sobre DUK GROUP. OJO: la planilla decía 100 gorras y la factura dice 940.'),
 ('DUK GROUP (DUK CATAMARCA SRL)', 'Bordado de gorras — COTIZACIÓN NO ADJUDICADA', 'Argentina',
  'Descartado', 'ARS', 2775360.0, 582825.6, 3358185.6, N, N, N, 2276.74, 'Transferencia',
  'DUK CATAMARCA S.R.L.', '30-71848443-6', 'Av. Libertador 1644 — Moreno CP 1744',
  'DUK CATAMARCA S.R.L.', 'Supervielle', '0270027410055637600021', 'DUKCATAMARCA.SUP', 'BSUPARBAXXX', '',
  'Quedó afuera: más caro que Remerasyestampados. Datos guardados por si se necesita.'),
 ('TEXTILRYU', '940 gorras × $2.367', 'Argentina', 'SIN INVOICE EMITIDA', 'ARS',
  2224980.0, N, N, N, N, N, N, 'Transferencia', '', '', '', '', '', '', '', 'BSCHARBAXXX', '',
  'REVISAR: 940 × $2.367 = $2.224.980 estimado. No hay invoice en la carpeta. ¿Se pagó por otra vía? Ojo con el doble 5% que se corrigió.'),
 ('Leotex', 'Lanyards — PAGO TOTAL', 'Argentina', 'Facturado', 'ARS',
  2070600.0, 434826.0, 2505426.0, 1.00, 2505426.0, 0.0, 1632.20, 'USDT BEP-20',
  'Leotex', '', 'Marengo 4178 — Villa Ballester CP 1653', 'Leotex', 'Galicia',
  '0070140820000004589444', '', 'GABAARBA', '2601018',
  'REVISAR: existen DOS invoices 2601018, una por el 50% ($1.252.713) y otra por el TOTAL ($2.505.426, ref. 2601013). Confirmar cuál se pagó.'),
 ('Derqui Impresiones', 'Gráfica Argentina', 'Argentina', 'Facturado', 'ARS',
  3025260.0, 635304.6, 3660564.6, 1.00, 3660564.6, 0.0, 2384.73, 'USDT BEP-20',
  'Germán Ariel Marino', '', 'Charrúa 3366 CABA — CP 1437', 'Germán Ariel Marino', 'BBVA',
  '0170107020000001722338', 'ESTADO.GRINGO.ENTERO', 'BBVAARBA', '2601014',
  'Contacto: Adrián — ventas@impresionesderqui.com.ar — 11 2456 7280.'),
 ('Derqui Impresiones', 'Gráfica Uruguay', 'Uruguay', 'Facturado', 'ARS',
  1564626.0, 328571.46, 1893197.46, 1.00, 1893197.46, 0.0, 1233.35, 'USDT BEP-20',
  'Germán Ariel Marino', '', 'Charrúa 3366 CABA — CP 1437', 'Germán Ariel Marino', 'BBVA',
  '0170107020000001722338', 'ESTADO.GRINGO.ENTERO', 'BBVAARBA', '2601015',
  'NO es de Argentina: es del evento de Uruguay.'),
 ('Agus (personal)', 'Handys Baofeng + diademas con VOX', 'Argentina', 'Facturado', 'USD',
  N, N, 560.0, 1.00, 560.0, 0.0, 560.0, 'USDT — Red BSC BEP-20', '', '', '', '', '',
  '0x7d0a6c347008305af1d6643c2382e1495af7dc0e', '', '', '',
  'Sin IVA. Ya comprados y pagados. Verificar wallet por segundo canal y guardar el hash de la transacción.'),
 ('FERROSVEL', 'Barras de hierro', 'Colombia', 'Facturado', 'COP',
  8732100.3, 1659099.06, 10391199.36, 1.00, 10391199.36, 0.0, 3403.07, 'Transferencia',
  '', '', '', '', '', '', '', '', '2601022',
  'OJO: es COLOMBIA, no Argentina. IVA 19%. Versión FINAL en la raíz de FACTURAS DE PAGO.'),
 ('Ignacio López', 'Reintegro de gastos Rep. Dominicana', 'Rep. Dominicana', 'Facturado', 'USD',
  N, N, 1075.98, 1.00, 1075.98, 0.0, 1075.98, 'Transferencia', '', '', '', '', '', '', '', '', '2601023',
  'Última invoice emitida. Próximo número disponible: 2601024.'),
 ('Carlos Calderón', 'Reintegro camerinos Panamá', 'Panamá', 'Facturado', 'USD',
  N, N, 334.98, 1.00, 334.98, 0.0, 334.98, 'Transferencia', '', '', '', '', '', '', '', '', '2601020',
  'Subtotal de 12 comprobantes USD 297,53 + comisión bancaria. Ver Informe de Gastos Panamá.'),
]

# ============================================================ diseño
import io, importlib.util, sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter as L

F, NAVY = 'Arial', '1F3864'
# cada estado a un semaforo: (relleno de fila, color de la pastilla, grupo)
EST = {
 'Facturado':                               ('E8F6F3', '117A65', 'Facturado / pagado'),
 'Facturado (50%)':                         ('E8F6F3', '117A65', 'Facturado / pagado'),
 'Cerrado + señado':                        ('E8F6F3', '117A65', 'Facturado / pagado'),
 'Cerrado — esperando seña':                ('FDF2E9', 'CA6F1E', 'Cerrado, falta señar'),
 'Cerrado — esperando presupuesto ajustado':('FDF2E9', 'CA6F1E', 'Cerrado, falta señar'),
 'Contratado':                              ('EBF5FB', '1F618D', 'En gestión'),
 'Negociación':                             ('EBF5FB', '1F618D', 'En gestión'),
 'Negociación — a recotizar':               ('EBF5FB', '1F618D', 'En gestión'),
 'Presupuesto recibido sin responder':      ('EBF5FB', '1F618D', 'En gestión'),
 'Esperando recotización':                  ('EBF5FB', '1F618D', 'En gestión'),
 'Cotización recibida':                     ('EBF5FB', '1F618D', 'En gestión'),
 'NUNCA LLEGÓ':                             ('FDEDEC', 'C0392B', 'Trabado / riesgo'),
 'Bloqueado':                               ('FDEDEC', 'C0392B', 'Trabado / riesgo'),
 'SIN INVOICE EMITIDA':                     ('FDEDEC', 'C0392B', 'Trabado / riesgo'),
 'Descartado':                              ('F2F4F4', '85929E', 'Descartado'),
}
ARS  = '"$"#,##0.00;("$"#,##0.00);-'
ARS0 = '"$"#,##0;("$"#,##0);-'
USD  = '"US$"#,##0.00;("US$"#,##0.00);-'
USD0 = '"US$"#,##0;("US$"#,##0);-'
PCT  = '0%'
BLANCO = Font(name=F, size=10, bold=True, color='FFFFFF')
NEGRO  = Font(name=F, size=10)
BOLD   = Font(name=F, size=10, bold=True)
GRIS   = Font(name=F, size=9, color='566573')
THIN   = Side(style='thin', color='D5D8DC')
BOX    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

wb = openpyxl.Workbook()

def encabezar(ws, titulo, bajada, cols, fila_hdr=4, congelar=3):
    ws.sheet_view.showGridLines = False
    nc = len(cols)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=nc)
    c = ws.cell(1, 1, titulo)
    c.font = Font(name=F, size=17, bold=True, color='FFFFFF')
    c.alignment = Alignment(vertical='center', indent=1)
    for j in range(1, nc + 1): ws.cell(1, j).fill = PatternFill('solid', fgColor=NAVY)
    ws.row_dimensions[1].height = 36
    c = ws.cell(2, 1, bajada); c.font = GRIS
    ws.row_dimensions[2].height = 18
    for j, (t, an, *_) in enumerate(cols, 1):
        c = ws.cell(fila_hdr, j, t); c.font = BLANCO
        c.fill = PatternFill('solid', fgColor=NAVY)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = BOX
        ws.column_dimensions[L(j)].width = an
    ws.row_dimensions[fila_hdr].height = 32
    ws.freeze_panes = ws.cell(fila_hdr + 1, congelar)
    return fila_hdr + 1

# ============================================================ 1. CONTROL
ws = wb.active; ws.title = 'CONTROL'
COLS = [('Proveedor', 30), ('Servicio / concepto', 44), ('Plaza', 15), ('Estado', 26), ('Moneda', 8),
        ('Neto', 15), ('IVA', 14), ('Total', 16), ('Seña %', 8), ('Monto seña', 16),
        ('Saldo pendiente', 16), ('Equiv. USD', 13), ('Vía de pago', 22),
        ('N.º invoice IIDAI', 13), ('Notas', 78)]
NC = len(COLS)
r = encabezar(ws, 'CONTROL DE PAGOS A PROVEEDORES · CMC 2026',
              'Actualizado al 3 de septiembre de 2026. Los datos fiscales y bancarios están en la hoja DATOS BANCARIOS.',
              COLS)
ini = r
for (prov, conc, plaza, estado, mon, neto, iva, total, spct, sena, saldo, usd,
     via, razon, cuit, dom, tit, banco, cbu, alias, swift, inv, nota) in FILAS:
    relleno, pastilla, _ = EST[estado]
    fmt  = ARS if mon in ('ARS', 'COP') else USD
    fmt0 = ARS0 if mon in ('ARS', 'COP') else USD0
    ws.cell(r, 1, prov).font = BOLD
    ws.cell(r, 2, conc).font = NEGRO
    ws.cell(r, 3, plaza).font = NEGRO
    c = ws.cell(r, 4, estado); c.font = Font(name=F, size=9, bold=True, color=pastilla)
    ws.cell(r, 5, mon or '—').font = GRIS
    for col, val in ((6, neto), (7, iva), (8, total), (10, sena), (11, saldo)):
        c = ws.cell(r, col, val)
        c.font = BOLD if col == 8 else NEGRO
        c.number_format = fmt
    c = ws.cell(r, 9, spct); c.number_format = PCT; c.font = NEGRO
    c = ws.cell(r, 12, usd); c.font = BOLD; c.number_format = USD
    ws.cell(r, 13, via or '—').font = GRIS
    ws.cell(r, 14, inv or '—').font = NEGRO
    ws.cell(r, 15, nota).font = GRIS
    for j in range(1, NC + 1):
        ws.cell(r, j).fill = PatternFill('solid', fgColor=relleno)
        ws.cell(r, j).border = BOX
        ws.cell(r, j).alignment = Alignment(wrap_text=(j in (2, 15)), vertical='center',
                                            horizontal='center' if j in (3, 5, 9, 14) else 'left',
                                            indent=0 if j in (3, 5, 9, 14) else 1)
    ws.row_dimensions[r].height = 40
    r += 1
fin = r - 1
# total: solo lo que esta en pesos, para no sumar monedas distintas
ws.cell(r, 2, 'TOTAL en pesos (sin la fila unificada de Ambient, que duplica)').font = Font(name=F, size=11, bold=True, color='FFFFFF')
c = ws.cell(r, 8, f'=SUMIFS($H${ini}:$H${fin},$E${ini}:$E${fin},"ARS",$A${ini}:$A${fin},"<>Grupo Ambient — TOTAL UNIFICADO")')
c.font = Font(name=F, size=12, bold=True, color='FFFFFF'); c.number_format = ARS0
c = ws.cell(r, 12, f'=SUM($L${ini}:$L${fin})')
c.font = Font(name=F, size=12, bold=True, color='FFFFFF'); c.number_format = USD0
for j in range(1, NC + 1):
    ws.cell(r, j).fill = PatternFill('solid', fgColor=NAVY); ws.cell(r, j).border = BOX
ws.row_dimensions[r].height = 28
ws.auto_filter.ref = f'A4:{L(NC)}{fin}'
# las filas sin monto cargado se marcan solas
ws.conditional_formatting.add(f'A{ini}:{L(NC)}{fin}',
    FormulaRule(formula=[f'AND($H{ini}="",$A{ini}<>"")'],
                font=Font(color='C0392B'), stopIfTrue=False))

# ============================================================ 2. DATOS BANCARIOS
ws2 = wb.create_sheet('DATOS BANCARIOS')
C2 = [('Proveedor', 30), ('Servicio / concepto', 40), ('Razón social', 30), ('CUIT / Tax ID', 18),
      ('Domicilio fiscal', 40), ('Titular de la cuenta', 28), ('Banco', 14),
      ('CBU / N.º de cuenta / wallet', 46), ('Alias', 22), ('SWIFT / BIC', 14), ('Invoice', 12)]
r2 = encabezar(ws2, 'DATOS FISCALES Y BANCARIOS',
               'Verificá siempre la cuenta o la wallet por un segundo canal antes de transferir.', C2)
n2 = 0
for (prov, conc, plaza, estado, mon, neto, iva, total, spct, sena, saldo, usd,
     via, razon, cuit, dom, tit, banco, cbu, alias, swift, inv, nota) in FILAS:
    if not any([razon, cuit, dom, tit, banco, cbu, alias, swift]):
        continue
    vals = [prov, conc, razon, cuit, dom, tit, banco, cbu, alias, swift, inv]
    for j, v in enumerate(vals, 1):
        c = ws2.cell(r2, j, v or '—')
        c.font = BOLD if j == 1 else (Font(name=F, size=9, color='1F3864') if j == 8 else NEGRO)
        c.border = BOX
        c.alignment = Alignment(wrap_text=(j in (2, 5, 8)), vertical='center', indent=1)
    relleno = 'FFFFFF' if n2 % 2 else 'F4F6F7'
    for j in range(1, len(C2) + 1): ws2.cell(r2, j).fill = PatternFill('solid', fgColor=relleno)
    ws2.row_dimensions[r2].height = 32
    r2 += 1; n2 += 1
ws2.auto_filter.ref = f'A4:{L(len(C2))}{r2-1}'

# ============================================================ 3. ALERTAS
ws3 = wb.create_sheet('ALERTAS')
C3 = [('Prioridad', 14), ('Proveedor', 30), ('Qué falta', 92), ('Estado', 26)]
r3 = encabezar(ws3, 'LO QUE ESTÁ TRABADO', 'Todo lo que no puede avanzar hasta que alguien haga algo.', C3, congelar=1)
ALERTAS = []
for f in FILAS:
    prov, conc, plaza, estado, mon, neto, iva, total = f[0], f[1], f[2], f[3], f[4], f[5], f[6], f[7]
    nota = f[22]
    grupo = EST[estado][2]
    if grupo == 'Trabado / riesgo':
        ALERTAS.append(('1 · TRABADO', prov, nota, estado))
    elif total is None and estado != 'Descartado':
        ALERTAS.append(('2 · FALTA MONTO', prov, nota, estado))
    elif 'REVISAR' in nota or 'Confirmar' in nota or 'VERIFICAR' in nota:
        ALERTAS.append(('3 · A VERIFICAR', prov, nota, estado))
ALERTAS.sort(key=lambda x: x[0])
for pri, prov, nota, estado in ALERTAS:
    color = {'1': 'FDEDEC', '2': 'FCF3CF', '3': 'EBF5FB'}[pri[0]]
    ws3.cell(r3, 1, pri).font = Font(name=F, size=9, bold=True,
             color={'1': 'C0392B', '2': '7D6608', '3': '1F618D'}[pri[0]])
    ws3.cell(r3, 2, prov).font = BOLD
    ws3.cell(r3, 3, nota).font = NEGRO
    ws3.cell(r3, 4, estado).font = GRIS
    for j in range(1, 5):
        ws3.cell(r3, j).fill = PatternFill('solid', fgColor=color)
        ws3.cell(r3, j).border = BOX
        ws3.cell(r3, j).alignment = Alignment(wrap_text=(j == 3), vertical='center', indent=1)
    ws3.row_dimensions[r3].height = 42
    r3 += 1

# ============================================================ 4. RESUMEN
ws4 = wb.create_sheet('RESUMEN')
C4 = [('Estado', 30), ('Ítems', 10), ('', 34), ('Total en pesos argentinos', 22), ('Equiv. USD', 16)]
r4 = encabezar(ws4, 'RESUMEN POR ESTADO', 'Solo pesos argentinos: los $10.391.199 de FERROSVEL son pesos colombianos y van aparte. No incluye la fila unificada de Ambient, que duplica los dos presupuestos de catering.', C4, congelar=1)
grupos = {}
for f in FILAS:
    g = EST[f[3]][2]
    grupos.setdefault(g, []).append(f)
ini4 = r4
for g in ['Trabado / riesgo', 'En gestión', 'Cerrado, falta señar', 'Facturado / pagado', 'Descartado']:
    fs = grupos.get(g, [])
    if not fs: continue
    # Solo ARS: los pesos colombianos de FERROSVEL no se pueden sumar con los argentinos.
    pesos = sum(f[7] or 0 for f in fs if f[4] == 'ARS'
                and f[0] != 'Grupo Ambient — TOTAL UNIFICADO')
    dolares = sum(f[11] or 0 for f in fs)
    color = {'Trabado / riesgo': 'FDEDEC', 'En gestión': 'EBF5FB', 'Cerrado, falta señar': 'FDF2E9',
             'Facturado / pagado': 'E8F6F3', 'Descartado': 'F2F4F4'}[g]
    ws4.cell(r4, 1, g).font = BOLD
    c = ws4.cell(r4, 2, len(fs)); c.font = NEGRO; c.alignment = Alignment(horizontal='center')
    c = ws4.cell(r4, 3, f'=REPT("|",B{r4}*3)'); c.font = Font(name=F, size=11, color=EST[fs[0][3]][1])
    c = ws4.cell(r4, 4, pesos or None); c.font = BOLD; c.number_format = ARS0
    c = ws4.cell(r4, 5, dolares or None); c.font = BOLD; c.number_format = USD0
    for j in range(1, 6):
        ws4.cell(r4, j).fill = PatternFill('solid', fgColor=color)
        ws4.cell(r4, j).border = BOX
        if j != 2:
            ws4.cell(r4, j).alignment = Alignment(vertical='center', indent=1)
    ws4.row_dimensions[r4].height = 24
    r4 += 1
ws4.cell(r4, 1, 'TOTAL').font = Font(name=F, size=12, bold=True, color='FFFFFF')
c = ws4.cell(r4, 2, f'=SUM(B{ini4}:B{r4-1})'); c.font = Font(name=F, size=12, bold=True, color='FFFFFF')
c.alignment = Alignment(horizontal='center')
c = ws4.cell(r4, 4, f'=SUM(D{ini4}:D{r4-1})'); c.font = Font(name=F, size=12, bold=True, color='FFFFFF'); c.number_format = ARS0
c = ws4.cell(r4, 5, f'=SUM(E{ini4}:E{r4-1})'); c.font = Font(name=F, size=12, bold=True, color='FFFFFF'); c.number_format = USD0
for j in range(1, 6):
    ws4.cell(r4, j).fill = PatternFill('solid', fgColor=NAVY); ws4.cell(r4, j).border = BOX
ws4.row_dimensions[r4].height = 28

r4 += 2
ws4.cell(r4, 1, 'CÓMO LEER LOS COLORES').font = Font(name=F, size=11, bold=True, color=NAVY); r4 += 1
for t, col in [('Rojo — trabado: nadie puede avanzar hasta que se resuelva.', 'FDEDEC'),
               ('Azul — en gestión: hay ida y vuelta con el proveedor.', 'EBF5FB'),
               ('Naranja — cerrado, falta señar: el número está, falta la plata.', 'FDF2E9'),
               ('Verde — facturado o pagado.', 'E8F6F3'),
               ('Gris — descartado, se guarda por si se necesita.', 'F2F4F4')]:
    c = ws4.cell(r4, 1, t); c.font = NEGRO; c.fill = PatternFill('solid', fgColor=col); c.border = BOX
    ws4.merge_cells(start_row=r4, start_column=1, end_row=r4, end_column=5)
    r4 += 1

wb.save('CONTROL_PAGOS_PROVEEDORES_CMC2026.xlsx')
print('filas de control:', fin - ini + 1)
print('datos bancarios:', n2)
print('alertas:', len(ALERTAS))
for g, fs in grupos.items():
    print('  {:<24} {:>2} items'.format(g, len(fs)))
