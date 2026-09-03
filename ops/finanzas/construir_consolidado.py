# -*- coding: utf-8 -*-
import json, datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as L
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule

HOY = datetime.date(2026, 9, 2)
D = json.load(open('datos_origen.json'))

FUENTE = 'Arial'
AZUL  = Font(name=FUENTE, size=10, color='0000FF')          # input manual
NEGRO = Font(name=FUENTE, size=10)
VERDE = Font(name=FUENTE, size=10, color='008000')          # link a otra hoja
TIT   = Font(name=FUENTE, size=14, bold=True, color='1F3864')
SUB   = Font(name=FUENTE, size=10, italic=True, color='595959')
HDR   = Font(name=FUENTE, size=10, bold=True, color='FFFFFF')
BOLD  = Font(name=FUENTE, size=10, bold=True)
FH    = PatternFill('solid', fgColor='1F3864')
FROJO = PatternFill('solid', fgColor='FADBD8')
FAMAR = PatternFill('solid', fgColor='FCF3CF')
FVERD = PatternFill('solid', fgColor='D5F5E3')
FGRIS = PatternFill('solid', fgColor='EAECEE')
ROJO_T, AMBAR_T, VERDE_T, GRIS_T = 'B03A2E', '7D6608', '196F3D', '566573'
FAMAR_IN = PatternFill('solid', fgColor='FFFF00')
THIN  = Side(style='thin', color='BFBFBF')
BOX   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MON   = '$#,##0.00;($#,##0.00);-'
MON0  = '$#,##0;($#,##0);-'
PCT   = '0.0%'
FECHA = 'DD/MM/YYYY'

wb = openpyxl.Workbook()

def hoja(nombre, titulo, subtitulo):
    ws = wb.create_sheet(nombre)
    ws.sheet_view.showGridLines = False
    ws['A1'] = titulo; ws['A1'].font = TIT
    ws['A2'] = subtitulo; ws['A2'].font = SUB
    return ws

def banner(ws, ncols):
    """Convierte el titulo de A1 en una banda azul de ancho completo.

    Se llama desde encabezados(), que es el unico lugar donde se conoce el
    ancho real de la hoja."""
    if ncols > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(1, 1)
    c.font = Font(name=FUENTE, size=16, bold=True, color='FFFFFF')
    c.alignment = Alignment(vertical='center', indent=1)
    for j in range(1, ncols + 1):
        ws.cell(1, j).fill = FH
    ws.row_dimensions[1].height = 34
    ws.row_dimensions[2].height = 18

def encabezados(ws, fila, cols):
    banner(ws, len(cols))
    for j, (txt, ancho) in enumerate(cols, 1):
        c = ws.cell(fila, j, txt); c.font = HDR; c.fill = FH
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = BOX
        ws.column_dimensions[L(j)].width = ancho
    ws.row_dimensions[fila].height = 30
    ws.freeze_panes = ws.cell(fila + 1, 1)

# ---------- orden canonico: bloque (categoria) y pais (itinerario) ----------
PAISES = ['Ecuador','Panamá','Rep. Dominicana','Colombia 1ra','México','Chile','Uruguay',
          'Argentina','Salvador','Colombia 2da','Costa Rica','Perú','Guatemala','España']
ORDEN_PAIS = {p: i for i, p in enumerate(PAISES)}
FECHA_EVENTO = {'Ecuador':'25 jul','Panamá':'08 ago','Rep. Dominicana':'15 ago','Colombia 1ra':'29 ago',
                'México':'05 sep','Chile':'12 sep','Uruguay':'19 sep','Argentina':'03 oct','Salvador':'10 oct',
                'Colombia 2da':'31 oct','Costa Rica':'07 nov','Perú':'14 nov','Guatemala':'21 nov','España':'28 nov'}
NORM_PAIS = {'Colombia':'Colombia 1ra','Colombia 1ra':'Colombia 1ra','Pendones':'Colombia 1ra',
             'Republica D.':'Rep. Dominicana','Rep. Dominicana':'Rep. Dominicana',
             'Panama':'Panamá','Panamá':'Panamá','Ecuador ':'Ecuador','Ecuador':'Ecuador',
             'México':'México','Mexico':'México','Chile':'Chile','Uruguay':'Uruguay','Argentina':'Argentina',
             'Salvador':'Salvador','Costa Rica':'Costa Rica','Perú':'Perú','Guatemala':'Guatemala','España':'España'}
def pais_canonico(p):
    return NORM_PAIS.get(str(p).strip(), str(p).strip())

BLOQUES = ['Lugar','Salas MM','Técnica','Producción','Dinámica','Merch','Catering','Trámites','Equipo','Otros']
ORDEN_BLOQUE = {b: i for i, b in enumerate(BLOQUES)}
NORM_BLOQUE = {'lugar':'Lugar','salas mm':'Salas MM','tecnica':'Técnica','técnica':'Técnica',
               'produccion':'Producción','producción':'Producción','dinamica':'Dinámica','dinámica':'Dinámica',
               'merch':'Merch','catering':'Catering','tramites':'Trámites','trámites':'Trámites','equipo':'Equipo'}
def bloque_canonico(c):
    return NORM_BLOQUE.get(str(c).strip().lower(), 'Otros')
def clave_orden(pais, cat):
    return (ORDEN_PAIS.get(pais_canonico(pais), 99), ORDEN_BLOQUE.get(bloque_canonico(cat), 99))

FBLOQUE = PatternFill('solid', fgColor='DCE6F1')   # banda de bloque (categoria)
FPAIS   = PatternFill('solid', fgColor='1F3864')   # banda de pais

def banda_pais(ws, fila, pais, ncols, extra=''):
    """Escribe la banda de encabezado de un pais y devuelve la fila siguiente."""
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=ncols)
    c = ws.cell(fila, 1, f'{pais}   ·   evento {FECHA_EVENTO.get(pais, "sin fecha")}' + (f'   ·   {extra}' if extra else ''))
    c.font = Font(name=FUENTE, size=12, bold=True, color='FFFFFF')
    c.fill = FPAIS
    c.alignment = Alignment(vertical='center', indent=1)
    ws.row_dimensions[fila].height = 24
    for j in range(1, ncols + 1):
        ws.cell(fila, j).fill = FPAIS
    return fila + 1

def banda_bloque(ws, fila, bloque, ncols):
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=ncols)
    c = ws.cell(fila, 1, f'      {bloque.upper()}')
    c.font = Font(name=FUENTE, size=9, bold=True, color='1F3864')
    c.alignment = Alignment(vertical='center')
    c.fill = FBLOQUE
    for j in range(1, ncols + 1):
        ws.cell(fila, j).fill = FBLOQUE
        ws.cell(fila, j).border = Border(top=Side(style='thin', color='C9A84C'))
    ws.row_dimensions[fila].height = 18
    return fila + 1

MAPP = {'Colombia 1ra':'Colombia','Pendones':'Colombia','Republica D.':'Rep. Dominicana',
        'Panama':'Panamá','Ecuador ':'Ecuador'}
kp = lambda p: MAPP.get(str(p).strip(), str(p).strip())
EV = {'Ecuador (25 Jul-26 Jul)':'Ecuador','Panama (09 ago-10 ago)':'Panamá',
      'Rep D. (15 ago-16 ago)':'Rep. Dominicana','Colombia (29 ago - 30 ago)':'Colombia',
      'Mexico (5 sep - 6 sep)':'México','Chile (12 sep-13 sep)':'Chile',
      'Uruguay (19 sep-20 sep)':'Uruguay','Costa Rica (1 ago-2 ago)':'Costa Rica',
      'Argentina (3 oct-4 oct)':'Argentina','Salvador (7 nov-8 nov)':'Salvador',
      'Guatemala (14 nov-15 nov)':'Guatemala','Perú ':'Perú','España (28 nov-29 nov)':'España'}

# ============================ 1. LEEME ============================
ws = hoja('LEEME', 'CUMBRE 2026 · Consolidado de gastos y pagos',
          'Unifica "Cronograma_Pagos_Cumbre.xlsx" y "INVERSIÓN GIRA CUMBRE 2026.xlsx" en una sola fuente de verdad.')
ws.column_dimensions['A'].width = 30; ws.column_dimensions['B'].width = 112
banner(ws, 2)
filas = [
 ('Fecha de corte', 'Todo el análisis está calculado al 02/09/2026 (celda B2 de la hoja TABLERO). Cambiala y se recalcula todo.'),
 ('Fuente 1 · CRONOGRAMA', 'Cronograma_Pagos_Cumbre (1).xlsx — Google Drive id 1MbTsbcMKau4WA4V2-0PM7EtJv2w4XA5P. Aporta el calendario de cuotas y el histórico de pagos.'),
 ('Fuente 2 · MASTER', 'INVERSIÓN GIRA CUMBRE 2026.xlsx — Google Drive id 1rRjt9YDSJKZR23nFlHJrAOf9CENc3CL2. Aporta contratos/reservas y el presupuesto línea por línea de cada evento.'),
 ('', ''),
 ('Cómo leer los colores', 'Azul = dato cargado a mano (editable). Negro = fórmula. Verde = link a otra hoja de este libro. Relleno amarillo = requiere decisión tuya.'),
 ('Rojo / Ámbar / Verde', 'Rojo = vencido o contradicción entre fuentes. Ámbar = vence dentro de 30 días o dato dudoso. Verde = conciliado.'),
 ('', ''),
 ('Supuesto de tipo de cambio', 'Las hojas por evento del MASTER están en moneda local con tasas inconsistentes. Ver hoja EVENTOS, columna "Nota FX": ahí está la tasa que se usó para cada conversión y por qué.'),
 ('Supuesto de no duplicación', 'CALENDARIO UNIFICADO (cuotas futuras y pagadas) y PAGOS REALIZADOS (histórico) son conjuntos disjuntos: el histórico termina el 03/07/2026 y el calendario empieza el 10/07/2026. Sumarlos NO duplica.'),
 ('Orden de las hojas de datos', 'CALENDARIO UNIFICADO, PAGOS REALIZADOS y CONTRATOS están agrupados por país en orden de gira (Ecuador → España) y, dentro de cada país, por bloque: Lugar, Salas MM, Técnica, Producción, Dinámica, Merch, Catering, Trámites, Equipo. Cada país cierra con su subtotal.'),
 ('Cómo reordenar', 'Las columnas País y Bloque están repetidas en cada fila de detalle y el filtro está activo en la fila de encabezado: podés filtrar por bloque para ver, por ejemplo, todos los pagos de Lugar de la gira. Las bandas azules son títulos, no datos: si ordenás la tabla se mezclan, mejor usá el filtro.'),
 ('', ''),
 ('Advertencia importante', 'El presupuesto de la hoja EVENTOS ($1.087.979) y el calendario de pagos ($417.765) NO son comparables: el primero es el costo total de producción de cada evento, el segundo sólo los contratos con anticipo/saldo. No los sumes.'),
 ('', ''),
 ('Qué falta cargar', 'Colombia 2da no tiene hoja de presupuesto en el MASTER. Perú, España y Colombia 2da no tienen ninguna cuota cargada en el calendario.'),
 ('Gastos de equipo en país', 'La hoja GASTOS EQUIPO está vacía a propósito: es el destino de lo que cargue el agente de WhatsApp. Hoy esos gastos no viven en ninguno de los dos archivos originales.'),
]
r = 4
for a, b in filas:
    ws.cell(r, 1, a).font = BOLD if a else NEGRO
    c = ws.cell(r, 2, b); c.font = NEGRO; c.alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[r].height = 28 if b else 8
    r += 1

# ============================ 2. TABLERO ============================
ws = hoja('TABLERO', 'Tablero por país', 'Presupuesto vs pagado vs pendiente. La columna "Brecha" mide cuánto se contradicen las dos fuentes.')
ws['A2'].font = SUB
ws['E2'] = 'Fecha de corte →'; ws['E2'].font = BOLD
ws['G2'] = HOY; ws['G2'].font = AZUL; ws['G2'].fill = FAMAR_IN; ws['G2'].number_format = FECHA
cols = [('País',18),('Presupuesto evento\n(MASTER, USD)',16),('Pagado según\nhoja del evento',16),
        ('Pagado según\nCRONOGRAMA',16),('Brecha entre\nfuentes',15),('Programado\n(pendiente)',15),
        ('VENCIDO\nal corte',14),('Fecha evento\nCRONOGRAMA',15),('Fecha evento\nMASTER',26),('Estado',34)]
encabezados(ws, 4, cols)

pres, pag_hoja, fx_nota, fecha_master = {}, {}, {}, {}
for e in D['eventos']:
    p = EV[e['hoja']]
    pres[p] = (e['total_local'] or 0) * e['factor']
    pag_hoja[p] = (e['pago_local'] or 0) * e['factor']
    fx_nota[p] = e['nota_fx']
    cab = e['cabecera'].replace('Check List Evento CUMBRE - ', '')
    fecha_master[p] = cab
FECHA_CRONO = {'Ecuador':'25/07/2026','Panamá':'08/08/2026','Rep. Dominicana':'15/08/2026',
  'Colombia':'29/08/2026','México':'05/09/2026','Chile':'12/09/2026','Uruguay':'19/09/2026',
  'Argentina':'03/10/2026','Salvador':'10/10/2026','Colombia 2da':'31/10/2026','Costa Rica':'07/11/2026',
  'Perú':'14/11/2026','Guatemala':'21/11/2026','España':'28/11/2026'}

hist, cal_pag, cal_prog, cal_venc = {}, {}, {}, {}
for h in D['historico']:
    hist[kp(h['pais'])] = hist.get(kp(h['pais']), 0) + h['monto']
for c in D['calendario']:
    p = kp(c['pais']); f = datetime.date.fromisoformat(c['fecha'])
    if c['estado'] == 'Pagado':
        cal_pag[p] = cal_pag.get(p, 0) + c['monto']
    else:
        cal_prog[p] = cal_prog.get(p, 0) + c['monto']
        if f < HOY: cal_venc[p] = cal_venc.get(p, 0) + c['monto']

orden = ['Ecuador','Panamá','Rep. Dominicana','Colombia','México','Chile','Uruguay','Argentina',
         'Salvador','Colombia 2da','Costa Rica','Perú','Guatemala','España']
r = 5
for p in orden:
    ws.cell(r,1,p).font = BOLD
    ws.cell(r,2,round(pres.get(p,0),2) or None).font = AZUL
    ws.cell(r,3,round(pag_hoja.get(p,0),2) or None).font = AZUL
    ws.cell(r,4,round(hist.get(p,0)+cal_pag.get(p,0),2) or None).font = AZUL
    ws.cell(r,5,f'=IFERROR(D{r}-C{r},"")').font = NEGRO
    ws.cell(r,6,round(cal_prog.get(p,0),2) or None).font = AZUL
    ws.cell(r,7,round(cal_venc.get(p,0),2) or None).font = AZUL
    ws.cell(r,8,FECHA_CRONO.get(p,'—')).font = AZUL
    ws.cell(r,9,fecha_master.get(p,'sin hoja en el MASTER')).font = AZUL
    est = []
    if cal_venc.get(p): est.append('VENCIDO')
    if abs(hist.get(p,0)+cal_pag.get(p,0)-pag_hoja.get(p,0)) > 500: est.append('fuentes no cierran')
    if p not in pres: est.append('sin presupuesto cargado')
    if not cal_prog.get(p) and not cal_pag.get(p) and not hist.get(p): est.append('sin pagos en el calendario')
    ws.cell(r,10,' · '.join(est) if est else 'conciliado').font = NEGRO
    for j in range(1,11):
        cc = ws.cell(r,j); cc.border = BOX
        if j in (2,3,4,5,6,7): cc.number_format = MON
    if cal_venc.get(p): ws.cell(r,7).fill = FROJO; ws.cell(r,10).fill = FROJO
    elif est and est != ['conciliado']: ws.cell(r,10).fill = FAMAR
    else: ws.cell(r,10).fill = FVERD
    r += 1
ws.cell(r,1,'TOTAL').font = BOLD
for j in range(2,8):
    c = ws.cell(r,j,f'=SUM({L(j)}5:{L(j)}{r-1})'); c.font = BOLD; c.number_format = MON; c.fill = FGRIS; c.border = BOX
ws.cell(r,1).fill = FGRIS; ws.cell(r,1).border = BOX
fila_total_tablero = r

r += 2
ws.cell(r,1,'Lectura rápida').font = BOLD
notas = [
 ('Pendiente real del calendario', f'=TABLERO!F{fila_total_tablero}', 'Las hojas resumen del CRONOGRAMA informan $417.765,32 porque suman también las 29 cuotas ya pagadas.'),
 ('Vencido al corte', f'=TABLERO!G{fila_total_tablero}', '10 cuotas en estado "Programado" con fecha anterior a la fecha de corte.'),
 ('Brecha entre las dos fuentes', f'=SUMPRODUCT(ABS(TABLERO!E5:E{fila_total_tablero-1}))', 'Suma de diferencias absolutas: cuánto dinero está registrado como pagado en un archivo y no en el otro.'),
]
for t, f, expl in notas:
    r += 1
    ws.cell(r,1,t).font = NEGRO
    c = ws.cell(r,2,f); c.font = VERDE; c.number_format = MON
    e = ws.cell(r,4,expl); e.font = SUB
ws.column_dimensions['D'].width = 16

# ============================ 3. CALENDARIO UNIFICADO ============================
ws = hoja('CALENDARIO UNIFICADO', 'Calendario unificado de cuotas',
          'Las 54 cuotas del CRONOGRAMA agrupadas por país (orden de gira) y, dentro de cada país, por bloque.')
COLS_CAL = [('Bloque',13),('Fecha pago',12),('Concepto / proveedor',46),('Cuota',20),
            ('Monto USD',13),('Estado',12),('Días vs corte',13),('Semáforo',13),
            ('País',15),('Contradice al MASTER',44)]
NC = len(COLS_CAL)
encabezados(ws, 4, COLS_CAL)
CONTRA = {
 ('Ecuador','Placas reconocimiento'): 'MASTER dice "Pendiente" (100% sin abonar)',
 ('Colombia 1ra','Camisetas'): 'MASTER dice "Pendiente" (100% sin abonar, total $433)',
 ('Colombia 1ra','100 Tarjetas platino'): 'MASTER dice "Pendiente" (100% sin abonar, total $611)',
 ('Argentina','Intercoms'): 'MASTER dice "Pendiente" (100%, total $560) y lo trata como pago único',
 ('Colombia 1ra','Gorras'): 'Etiqueta "Cuota 1 de 1 (100%)" pero es la 1ra mitad del saldo de $2.000',
 ('Panamá','Faranda (Salas MM)'): 'La hoja Panamá registra $1.190 en Faranda + $1.375 en Hotel America Golden Tower',
}
cal_orden = sorted(D['calendario'],
                   key=lambda c: (clave_orden(c['pais'], c['cat']), c['fecha']))
r = 5
filas_detalle_cal, filas_subtotal_pais = [], []
pais_actual = bloque_actual = None
inicio_pais = None
def cerrar_pais(ws, r, inicio, pais):
    """Escribe la fila de subtotal del país y devuelve la fila siguiente."""
    ws.cell(r,1,f'Subtotal {pais}').font = BOLD
    for j, col in ((5,'E'),):
        pass
    c = ws.cell(r,5,f'=SUMIF($F${inicio}:$F${r-1},"<>",$E${inicio}:$E${r-1})')
    c.font = BOLD; c.number_format = MON
    c = ws.cell(r,6,f'=SUMIF($F${inicio}:$F${r-1},"Programado",$E${inicio}:$E${r-1})')
    c.font = BOLD; c.number_format = MON
    ws.cell(r,7,'pendiente →').font = SUB
    for j in range(1, NC+1):
        ws.cell(r,j).fill = FGRIS; ws.cell(r,j).border = BOX
    return r + 1

for c in cal_orden:
    p = pais_canonico(c['pais']); b = bloque_canonico(c['cat'])
    if p != pais_actual:
        if pais_actual is not None:
            r = cerrar_pais(ws, r, inicio_pais, pais_actual); filas_subtotal_pais.append(r-1)
        r = banda_pais(ws, r, p, NC)
        inicio_pais = r
        pais_actual, bloque_actual = p, None
    if b != bloque_actual:
        r = banda_bloque(ws, r, b, NC); bloque_actual = b
    f = datetime.date.fromisoformat(c['fecha'])
    ws.cell(r,1,b).font = SUB
    ws.cell(r,2,f).number_format = FECHA; ws.cell(r,2).font = AZUL
    ws.cell(r,3,c['concepto']).font = AZUL
    ws.cell(r,4,c['cuota']).font = AZUL
    cc = ws.cell(r,5,c['monto']); cc.number_format = MON; cc.font = AZUL
    ws.cell(r,6,c['estado']).font = AZUL
    ws.cell(r,7,f'=IF(F{r}="Pagado","",TABLERO!$G$2-B{r})').font = NEGRO
    ws.cell(r,8,f'=IF(F{r}="Pagado","PAGADO",IF(B{r}<TABLERO!$G$2,"VENCIDO",IF(B{r}<=TABLERO!$G$2+30,"< 30 dias","futuro")))').font = NEGRO
    ws.cell(r,9,p).font = NEGRO
    txt = ''
    for (pp, con), msg in CONTRA.items():
        if p == pp and con.lower() in c['concepto'].lower(): txt = msg
    ws.cell(r,10,txt).font = NEGRO
    if c['estado'] != 'Pagado' and f < HOY:
        for j in range(1,NC+1): ws.cell(r,j).fill = FROJO
    elif c['estado'] != 'Pagado' and (f-HOY).days <= 30:
        for j in range(1,NC+1): ws.cell(r,j).fill = FAMAR
    if txt: ws.cell(r,10).fill = FAMAR
    for j in range(1,NC+1): ws.cell(r,j).border = BOX
    filas_detalle_cal.append(r)
    r += 1
r = cerrar_pais(ws, r, inicio_pais, pais_actual); filas_subtotal_pais.append(r-1)

r += 1
suma_sub = '+'.join(f'E{x}' for x in filas_subtotal_pais)
suma_prog = '+'.join(f'F{x}' for x in filas_subtotal_pais)
fin_detalle = r - 2      # ultima fila de detalle, antes del bloque de totales
for etiqueta, formula, relleno in [
    ('TOTAL del calendario', f'={suma_sub}', FGRIS),
    ('Pagado',  f'=SUMIF($F$5:$F${r-2},"Pagado",$E$5:$E${r-2})', FVERD),
    ('Pendiente real', f'={suma_prog}', FAMAR),
    ('De eso, VENCIDO', f'=SUMIFS($E$5:$E${r-2},$F$5:$F${r-2},"Programado",$B$5:$B${r-2},"<"&TABLERO!$G$2)', FROJO)]:
    ws.cell(r,3,etiqueta).font = Font(name=FUENTE, size=11, bold=True)
    c = ws.cell(r,5,formula)
    c.font = Font(name=FUENTE, size=13, bold=True); c.number_format = MON0
    for j in range(3, 6):
        ws.cell(r,j).fill = relleno; ws.cell(r,j).border = BOX
        ws.cell(r,j).alignment = Alignment(vertical='center', indent=1)
    ws.row_dimensions[r].height = 24
    r += 1
ws.auto_filter.ref = f'A4:{L(NC)}4'

# Semaforo con color: el estado de cada cuota se lee de un vistazo.
for texto, relleno, color in [('VENCIDO', 'FADBD8', ROJO_T), ('< 30 dias', 'FCF3CF', AMBAR_T),
                              ('PAGADO', 'D5F5E3', VERDE_T), ('futuro', 'F2F4F4', GRIS_T)]:
    ws.conditional_formatting.add(
        f'H5:H{fin_detalle}',
        FormulaRule(formula=[f'$H5="{texto}"'], fill=PatternFill(bgColor=relleno),
                    font=Font(color=color, bold=True)))

# ============================ 4. PAGOS REALIZADOS ============================
ws = hoja('PAGOS REALIZADOS', 'Histórico de pagos ejecutados (marzo–julio 2026)',
          'Anticipos previos al arranque del calendario, agrupados por país y bloque. Conjunto disjunto del CALENDARIO UNIFICADO.')
COLS_H = [('Bloque',13),('Fecha',12),('Proveedor / concepto',52),('% del total',12),
          ('Monto USD',14),('País',16),('País en el archivo original',24)]
NC = len(COLS_H)
encabezados(ws, 4, COLS_H)
hist_orden = sorted(D['historico'], key=lambda h: (clave_orden(h['pais'], h['cat']), h['fecha']))
r = 5
subtot_h = []
pais_actual = bloque_actual = None; inicio_pais = None
for h in hist_orden:
    p = pais_canonico(h['pais']); b = bloque_canonico(h['cat'])
    if p != pais_actual:
        if pais_actual is not None:
            ws.cell(r,3,f'Subtotal {pais_actual}').font = BOLD
            c = ws.cell(r,5,f'=SUM(E{inicio_pais}:E{r-1})'); c.font = BOLD; c.number_format = MON
            for j in range(1,NC+1): ws.cell(r,j).fill = FGRIS; ws.cell(r,j).border = BOX
            subtot_h.append(r); r += 1
        r = banda_pais(ws, r, p, NC); inicio_pais = r
        pais_actual, bloque_actual = p, None
    if b != bloque_actual:
        r = banda_bloque(ws, r, b, NC); bloque_actual = b
    ws.cell(r,1,b).font = SUB
    ws.cell(r,2,datetime.date.fromisoformat(h['fecha'])).number_format = FECHA
    ws.cell(r,2).font = AZUL
    ws.cell(r,3,h['concepto']).font = AZUL
    ws.cell(r,4,h['pct']).font = AZUL
    c = ws.cell(r,5,h['monto']); c.number_format = MON; c.font = AZUL
    ws.cell(r,6,p).font = NEGRO
    orig = str(h['pais']).strip()
    ws.cell(r,7,orig if orig != p else '').font = SUB
    if orig != p: ws.cell(r,7).fill = FAMAR
    if str(h['pct']).strip() == '?': ws.cell(r,4).fill = FAMAR
    for j in range(1,NC+1): ws.cell(r,j).border = BOX
    r += 1
ws.cell(r,3,f'Subtotal {pais_actual}').font = BOLD
c = ws.cell(r,5,f'=SUM(E{inicio_pais}:E{r-1})'); c.font = BOLD; c.number_format = MON
for j in range(1,NC+1): ws.cell(r,j).fill = FGRIS; ws.cell(r,j).border = BOX
subtot_h.append(r); r += 2
ws.cell(r,3,'TOTAL histórico').font = BOLD
c = ws.cell(r,5,'=' + '+'.join(f'E{x}' for x in subtot_h)); c.font = BOLD; c.number_format = MON; c.fill = FGRIS; c.border = BOX
ws.auto_filter.ref = f'A4:{L(NC)}4'

# ============================ 5. CONTRATOS ============================
ws = hoja('CONTRATOS', 'Contratos y reservas (hoja PAGOS del MASTER)',
          '44 contratos por país y bloque. "Saldo master" es el saldo escrito el día de la reserva: no se actualiza solo.')
COLS_C = [('Bloque',13),('Fecha reserva',13),('Proveedor',42),('Valor total',14),('% 1er pago',11),
          ('1er pago',14),('Saldo master',14),('Fecha saldo',13),('País',16),
          ('Fila origen',10),('Alerta',50)]
NC = len(COLS_C)
encabezados(ws, 4, COLS_C)
cont_orden = sorted(D['contratos'], key=lambda c: (clave_orden(c['pais'], c['cat']), c['proveedor'].lower()))
r = 5
subtot_c = []
pais_actual = bloque_actual = None; inicio_pais = None
def cerrar_contratos(ws, r, inicio, pais):
    ws.cell(r,3,f'Subtotal {pais}').font = BOLD
    for j in (4,7):
        c = ws.cell(r,j,f'=SUM({L(j)}{inicio}:{L(j)}{r-1})'); c.font = BOLD; c.number_format = MON
    for j in range(1,NC+1): ws.cell(r,j).fill = FGRIS; ws.cell(r,j).border = BOX
    return r + 1
for c in cont_orden:
    p = pais_canonico(c['pais']); b = bloque_canonico(c['cat'])
    if p != pais_actual:
        if pais_actual is not None:
            r = cerrar_contratos(ws, r, inicio_pais, pais_actual); subtot_c.append(r-1)
        r = banda_pais(ws, r, p, NC); inicio_pais = r
        pais_actual, bloque_actual = p, None
    if b != bloque_actual:
        r = banda_bloque(ws, r, b, NC); bloque_actual = b
    ws.cell(r,1,b).font = SUB
    if c['fecha_reserva']:
        ws.cell(r,2,datetime.date.fromisoformat(c['fecha_reserva'])).number_format = FECHA
    ws.cell(r,2).font = AZUL
    ws.cell(r,3,c['proveedor']).font = AZUL
    for j, v in ((4,c['total']),(5,c['pct1']),
                 (6,c['pago1'] if c['pago1'] is not None else c['pago1_txt']),(7,c['saldo'])):
        cc = ws.cell(r,j,v); cc.font = AZUL
        if j in (4,6,7) and isinstance(v,(int,float)): cc.number_format = MON
        if j == 5 and isinstance(v,(int,float)): cc.number_format = PCT
    if c['fecha_saldo']:
        ws.cell(r,8,datetime.date.fromisoformat(c['fecha_saldo'])).number_format = FECHA
    ws.cell(r,8).font = AZUL
    ws.cell(r,9,p).font = NEGRO
    ws.cell(r,10,c['fila']).font = SUB
    al = []
    if c['pago1_txt']: al.append('Sin abonar según MASTER — revisar contra el CALENDARIO')
    if c['pago1'] is None and not c['pago1_txt'] and c['total']: al.append('Contrato sin ningún dato de pago en el MASTER')
    if str(c['pct1']).strip() == '?': al.append('% del primer pago desconocido')
    ws.cell(r,11,' · '.join(al)).font = NEGRO
    if al: ws.cell(r,11).fill = FAMAR
    for j in range(1,NC+1): ws.cell(r,j).border = BOX
    r += 1
r = cerrar_contratos(ws, r, inicio_pais, pais_actual); subtot_c.append(r-1)
r += 1
ws.cell(r,3,'TOTAL contratado').font = BOLD
c = ws.cell(r,4,'=' + '+'.join(f'D{x}' for x in subtot_c)); c.font = BOLD; c.number_format = MON; c.fill = FGRIS; c.border = BOX
c = ws.cell(r,7,'=' + '+'.join(f'G{x}' for x in subtot_c)); c.font = BOLD; c.number_format = MON; c.fill = FGRIS; c.border = BOX
ws.auto_filter.ref = f'A4:{L(NC)}4'

# ============================ 6. EVENTOS ============================
ws = hoja('EVENTOS', 'Presupuesto por evento y tipos de cambio',
          'Cada hoja del MASTER está en su moneda. Acá se documenta la tasa usada y por qué.')
encabezados(ws, 4, [('País',17),('Hoja origen',26),('Moneda',9),('Total en moneda local',18),
                    ('Tasa a USD',12),('Total USD',14),('Pagado local',16),('Pagado USD',14),
                    ('Nota FX',52),('Fecha en el MASTER',40)])
r = 5
for e in D['eventos']:
    p = EV[e['hoja']]
    ws.cell(r,1,p).font = BOLD
    ws.cell(r,2,e['hoja'].strip()).font = AZUL
    ws.cell(r,3,e['moneda']).font = AZUL
    ws.cell(r,4,e['total_local']).font = AZUL; ws.cell(r,4).number_format = '#,##0.00'
    ws.cell(r,5,e['factor']).font = AZUL; ws.cell(r,5).number_format = '0.000000'
    ws.cell(r,6,f'=D{r}*E{r}').font = NEGRO; ws.cell(r,6).number_format = MON
    ws.cell(r,7,e['pago_local']).font = AZUL; ws.cell(r,7).number_format = '#,##0.00'
    ws.cell(r,8,f'=IF(G{r}="","",G{r}*E{r})').font = NEGRO; ws.cell(r,8).number_format = MON
    ws.cell(r,9,e['nota_fx']).font = NEGRO
    ws.cell(r,10,e['cabecera'].replace('Check List Evento CUMBRE - ','')).font = AZUL
    if e['moneda'] != 'USD' or 'heredada' in e['nota_fx'] or 'vacia' in e['nota_fx'] or 'Sin celda' in e['nota_fx']:
        ws.cell(r,9).fill = FAMAR
    for j in range(1,11): ws.cell(r,j).border = BOX
    r += 1
ws.cell(r,1,'TOTAL').font = BOLD
for j in (6,8):
    c = ws.cell(r,j,f'=SUM({L(j)}5:{L(j)}{r-1})'); c.font = BOLD; c.number_format = MON; c.fill = FGRIS
ws.cell(r+2,1,'Colombia 2da (31/10/2026) no tiene hoja de presupuesto en el MASTER: su costo no está dentro de este total.').font = SUB

# ============================ 7. HALLAZGOS ============================
ws = hoja('HALLAZGOS', 'Hallazgos de la conciliación', 'Ordenados por impacto. La columna "Acción" es lo que hay que decidir.')
encabezados(ws, 4, [('#',5),('Severidad',12),('Hallazgo',54),('Evidencia',72),('Impacto USD',14),('Acción',52)])
H = [
 ('ALTA','El informe quincenal sobrestima el pendiente','Las hojas PRESUPUESTO QUINCENAL y VISTA POR EVENTO suman las 54 filas del calendario, incluidas las 29 ya marcadas "Pagado". Informan $417.765,32 de pendiente.',188854.32,'Reemplazar el informe a finanzas por la hoja TABLERO. Pendiente real: $228.911,00.'),
 ('ALTA','Cuotas vencidas sin pagar','10 cuotas en estado "Programado" con fecha anterior al 02/09/2026. La más vieja lleva 54 días (Colombia merch, 10/07).',106250.50,'Confirmar con Erick cuáles se pagaron y no se marcaron, y cuáles están realmente en mora.'),
 ('ALTA','Las dos fuentes no cierran en pagos','Suma de diferencias absolutas por país entre "pagado según la hoja del evento" y "pagado según el cronograma".',275818.32,'Definir cuál archivo es la fuente de verdad de PAGOS. Propuesta: el consolidado, alimentado por el agente.'),
 ('ALTA','Colombia: $86.480 pagados que la hoja del evento reporta como $0','City Hall 2 x $36.500 + merch figuran "Pagado" en el cronograma. Las 64 líneas de la hoja Colombia tienen la columna Pago en 0.',86480.00,'Cargar los pagos en la hoja de Colombia o migrar a este consolidado.'),
 ('ALTA','Fechas de evento en conflicto entre archivos','Colombia: 29/08 vs 26-27/09 · México: 05/09 vs 01-02/11 · Costa Rica: 07/11 vs 01-02/08 · Salvador: 10/10 vs 07-08/11 · Guatemala: 21/11 vs 14-15/11 · Panamá: 08/08 vs 09-10/08.',0,'Fijar el calendario real de la gira. Toda la regla de pagos ("Lugar: 10 días antes") depende de esta fecha.'),
 ('ALTA','14 contratos sin ningún dato de pago en el MASTER','City Hall $73.000, Crowne Plaza $34.000, Plataforma AV Panamá y RD $15.000 c/u, Faranda $2.000, Wyndham Quintas $1.706, Varillas, Impresos, Unifilas, Ambulancia, Bolsas.',150382.00,'La única traza de esos pagos es el cronograma. Adjuntar comprobantes antes del cierre contable.'),
 ('MEDIA','Salvador tiene dos sedes vivas','Contrato pagado y programado: Círculo Militar ($24.589, $7.376 abonado). La hoja del evento presupuesta Hotel Hilton ($31.000 sala + $7.890 técnica).',31000.00,'Confirmar sede con Aleja. Si se cambió, reclamar o reasignar el anticipo del Círculo Militar.'),
 ('MEDIA','Panamá: sala Mastermind pagada en dos hoteles','El cronograma pagó Faranda "Salas MM" $2.000 en dos cuotas. La hoja Panamá registra $1.190 en Faranda y $1.375 en Hotel America Golden Tower, ambos pagados.',810.00,'Reclamar el crédito no usado en Faranda (~$810).'),
 ('MEDIA','Contradicciones de estado en 5 conceptos','Placas reconocimiento EC $390, Camisetas CO $433, Tarjetas platino CO $611, Intercoms AR $560: el cronograma los da por pagados o parcialmente pagados, el MASTER dice "Pendiente".',1994.00,'Verificar contra extractos bancarios y unificar el estado en este consolidado.'),
 ('MEDIA','Ecuador cerró con saldo sin liquidar','Evento del 25-26/07. Quinta by Wyndham: hospedaje $3.110 y salas MM/empresarios $2.082 sin pagar. Vuelos LATAM cruzados: $1.564 imputados a un sobrepeso de $260 y sólo $299 de $2.917 de vuelos.',6506.07,'Liquidar con el hotel y corregir la imputación de LATAM.'),
 ('MEDIA','Ecuador: catering $5.074 nunca programado','El MASTER lo marca "Pendiente"; el cronograma no tiene ninguna fila para el catering de Quorum Quito. El evento ya ocurrió.',5074.00,'Confirmar si se pagó. Si no, es una deuda vencida hace 5 semanas.'),
 ('MEDIA','Guatemala: presupuesto 3x el contrato','Contrato Parque de la Industria $5.000 ($1.500 abonado, saldo $3.500 en 2 cuotas). La hoja del evento presupuesta $14.351 de ítems del mismo venue.',9351.00,'Pedir cotización firme a Michelle Taylor antes de fijar las cuotas de octubre y noviembre.'),
 ('MEDIA','Tres eventos sin ninguna cuota cargada','Colombia 2da (31/10), Perú (14/11) y España (28/11) figuran con "-" en el cronograma. Perú y España tienen $51.504 y $51.228 de presupuesto; Colombia 2da no tiene ni hoja.',102732.00,'Cargar contratos y fechas de pago. Son los tres últimos eventos del año y hoy no tienen plan financiero.'),
 ('MEDIA','Tipos de cambio incoherentes dentro de la misma hoja','México usa 0,059 para presupuestar y 0,051 para pagos. Colombia usa 3.700 y 4.080. Chile y Uruguay tienen la celda de tasa en 0. Salvador, Guatemala y Perú apuntan al link de "peso mexicano a dólar" heredado por copiar la hoja.',0,'Fijar una tasa por evento en la hoja EVENTOS y usarla en todas las columnas.'),
 ('BAJA','El "Gral Resume" del MASTER está roto','La columna PAGADO da #REF! en Guatemala y Perú, $0 en Panamá, Colombia, Argentina, Uruguay y España, y $3,48 en Chile.',0,'Dejar de usarlo. La hoja TABLERO lo reemplaza.'),
 ('BAJA','Etiquetas de cuota inconsistentes','Colombia Gorras: una fila dice "Cuota 1 de 1 (100%)" por $1.000 y otra "Cuota 2 de 2 (50%)" por $1.000, cuando son las dos mitades del mismo saldo de $2.000.',0,'Normalizar el texto de cuota. El agente lo va a generar automáticamente.'),
]
r = 5
for i, (sev, tit, ev, imp, acc) in enumerate(H, 1):
    ws.cell(r,1,i).font = NEGRO
    c = ws.cell(r,2,sev); c.font = BOLD
    c.fill = FROJO if sev == 'ALTA' else (FAMAR if sev == 'MEDIA' else FGRIS)
    ws.cell(r,3,tit).font = BOLD
    ws.cell(r,4,ev).font = NEGRO
    cc = ws.cell(r,5,imp or None); cc.number_format = MON; cc.font = NEGRO
    ws.cell(r,6,acc).font = NEGRO
    for j in range(1,7):
        ws.cell(r,j).border = BOX
        ws.cell(r,j).alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[r].height = 46
    r += 1

# ============================ 8. GASTOS EQUIPO ============================
ws = hoja('GASTOS EQUIPO', 'Gastos del equipo en país',
          'Hoja destino del agente de WhatsApp. Hoy estos gastos no existen en ninguno de los dos archivos originales.')
encabezados(ws, 4, [('ID',12),('Fecha',12),('País / evento',16),('Persona',26),('Rol',14),
                    ('Comercio / proveedor',28),('Concepto',34),('Categoría',16),('Moneda',9),
                    ('Monto local',13),('Tasa',10),('Monto USD',13),('Comprobante',30),
                    ('Estado',14),('N° factura reintegro',20)])
dv = DataValidation(type='list', formula1='"Camerinos,Catering,Transporte,Hospedaje,Insumos,Impresiones,Otros"', allow_blank=True)
ws.add_data_validation(dv); dv.add('H5:H500')
dv2 = DataValidation(type='list', formula1='"Cargado,Aprobado,Facturado,Reintegrado"', allow_blank=True)
ws.add_data_validation(dv2); dv2.add('N5:N500')
ejemplo = ['GE-2026-0001', datetime.date(2026,8,9), 'Panamá', 'Carlos Andrés Calderón Arbeláez', 'Staff',
           'Riba Smith', 'Camerinos speakers / staff', 'Camerinos', 'USD', 133.81, 1.0, None,
           'foto ticket WhatsApp', 'Facturado', '2601020']
for j, v in enumerate(ejemplo, 1):
    c = ws.cell(5,j,v); c.font = AZUL; c.fill = FAMAR_IN
    if j == 2: c.number_format = FECHA
    if j in (10,12): c.number_format = MON
ws.cell(5,12,'=J5*K5').font = NEGRO
for r in range(6, 60):
    ws.cell(r,12,f'=IF(J{r}="","",J{r}*K{r})').font = NEGRO
    ws.cell(r,12).number_format = MON
    for j in range(1,16): ws.cell(r,j).border = BOX
ws.cell(3,1,'Fila 5 = ejemplo con formato esperado (tomado de la factura 2601020 de Panamá). Borrala cuando el agente empiece a escribir.').font = SUB

del wb['Sheet']
wb.save('CUMBRE_2026_Consolidado.xlsx')
print('OK')
