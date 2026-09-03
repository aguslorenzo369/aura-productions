# -*- coding: utf-8 -*-
"""Parte ARGENTINA_CMC2026_Costo_Total.xlsx en tres libros más chicos.

Existe porque el libro entero (42 KB) no se puede subir por la conexión MCP de
Drive: los binarios hay que transcribirlos completos en un solo envío y a ese
tamaño se cortan. Partido en tres, cada archivo entra.

El corte respeta las dependencias entre hojas. TABLERO y COSTOS UNIFICADOS se
referencian mutuamente y PAGOS lee de COSTOS UNIFICADOS, así que esas cuatro
hojas viajan juntas. El resto sólo usa TABLERO!$C$4 (el dólar operativo), que
se reemplaza por el literal 1510 para que cada archivo quede autosuficiente:
sin eso, las fórmulas darían #REF y los totales aparecerían en cero.
"""
import openpyxl, re, os

ORIGEN = 'ARGENTINA_CMC2026_Costo_Total.xlsx'
TC = 1510

# LEEME viaja en la parte 2 y no en la 1 sólo por tamaño: la 1 ya carga la tabla
# de costos, que es la hoja más pesada, y hay que dejarla lo más liviana posible.
PARTES = [
    ('ARGENTINA CMC 2026 · 1 de 3 — Costos y pagos.xlsx',
     ['COSTOS UNIFICADOS', 'TABLERO', 'PAGOS']),
    ('ARGENTINA CMC 2026 · 2 de 3 — Cómo leerlo, merch y decisiones.xlsx',
     ['LEEME', 'MERCH', 'DECISIONES']),
    ('ARGENTINA CMC 2026 · 3 de 3 — Ahorros y cotizaciones.xlsx',
     ['AHORROS', 'COTIZACIONES TÉCNICA', 'INFRAESTRUCTURA']),
]

def desreferenciar(ws):
    """TABLERO!$C$4 -> 1510, para las hojas que viajan sin TABLERO."""
    n = 0
    for fila in ws.iter_rows():
        for celda in fila:
            v = celda.value
            if isinstance(v, str) and 'TABLERO!' in v:
                celda.value = re.sub(r"TABLERO!\$?C\$?4", str(TC), v)
                n += 1
    return n

for nombre, hojas in PARTES:
    wb = openpyxl.load_workbook(ORIGEN)
    for h in list(wb.sheetnames):
        if h not in hojas:
            del wb[h]
    # dejar las hojas en el orden pedido
    wb._sheets = [wb[h] for h in hojas]
    tocadas = 0 if 'TABLERO' in hojas else sum(desreferenciar(wb[h]) for h in hojas)
    wb.save(nombre)
    kb = os.path.getsize(nombre) / 1024
    print('{:6.1f} KB  {:<52} {} · {} fórmulas desreferenciadas'.format(
        kb, nombre, ', '.join(hojas), tocadas))
