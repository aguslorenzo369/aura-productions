# -*- coding: utf-8 -*-
"""Checklist de pagos de Argentina · CMC 2026.

Emite dos cosas desde la misma fuente de datos:
  · checklist_pagos_argentina.xlsx  — la planilla formateada, para subir a Sheets
  · checklist_pagos_argentina.csv   — la misma tabla en texto plano

Los montos salen de construir_argentina.py, no se transcriben a mano, y el
total del checklist tiene que dar exactamente lo que falta desembolsar.
"""
import importlib.util, sys, csv, io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter as L

spec = importlib.util.spec_from_file_location('gen', 'construir_argentina.py')
m = importlib.util.module_from_spec(spec); sys.modules['gen'] = m; spec.loader.exec_module(m)
TC = m.TC_DEFECTO

def pct(it): return it[10] if len(it) > 10 else 1.0
def usd(it):
    _, ru, pr, de, mon, monto, iva, est, e, fu = it[:10]
    if est in ('Bonificado', 'Reemplazado'): return 0.0
    if monto:
        b = monto * (1 + iva)
        return pct(it) * (b if mon == 'USD' else b / TC)
    return pct(it) * e
TOTAL = sum(usd(i) for i in m.ITEMS if i[7] != 'Alternativa')
PAGADO = 55891.15
FALTA = TOTAL - PAGADO

MET   = 'Grupo MET · Lalo Aizenberg'
RURAL = 'La Rural'
# (tramo, rubro, proveedor, concepto, a_pagar, vence, como, contacto, telefono, nota)
FILAS = [
 ('1 · VENCIDO O YA FACTURADO', 'Sede', RURAL, 'Saldo del contrato CC-00957, cuota 1 de 2',
  15777.00, '24 ago — VENCIDA', 'Transferencia', 'Victoria Grosi · vgrosi@larural.com.ar',
  'congresosyeventos@larural.com.ar', 'Lleva 10 días vencida.'),
 ('1 · VENCIDO O YA FACTURADO', 'Merch', 'REMERASYESTAMPADOS', 'Remeras premium x250 — factura 2601017',
  2780.64, 'facturado 16 ago', 'USDT BEP20', 'Leandro Petracci · CUIT 20-30367082-4',
  'Uruguay 41, Villa Martelli', 'Wallet 0xae874c3db52ca1e45604549a5527f450a877b97a'),
 ('1 · VENCIDO O YA FACTURADO', 'Merch', 'REMERASYESTAMPADOS', 'Bordado de 940 gorras — factura 2601016',
  1721.95, 'facturado 10 ago', 'USDT BEP20', 'Leandro Petracci · CUIT 20-30367082-4',
  'Uruguay 41, Villa Martelli', 'Mismo wallet que las remeras.'),

 ('2 · ANTICIPOS QUE DESTRABAN CONTRATOS', 'Técnica', MET, 'Producción técnica — resto del anticipo del 50%',
  25000.00, 'urgente', 'Transferencia / Mercury', 'Lalo Aizenberg', '11 6411-6748',
  'El 50% son US$30.000; ya se pagaron US$5.000 con la factura INV-11 del 31/07.'),
 ('2 · ANTICIPOS QUE DESTRABAN CONTRATOS', 'Catering', 'Grupo Ambient · Ángeles', 'Catering — anticipo del 50%',
  7284.50, 'urgente', 'Transferencia', 'Ángeles', 'angeles@grupoambient.com.ar',
  'Total cerrado US$14.569 por las dos propuestas.'),

 ('3 · PROVEEDORES DEL PREDIO — firmar y señar', 'Mobiliario', 'FDL Eventos · Federico', '5.500 sillas + flete',
  20674.17, 'a definir', 'A resolver', 'Federico', '11 4403-0698 / 11 4040-7012',
  'PENDIENTE: no confirmaron si pueden facturar al exterior. Lo hacen con un tercero que emite en EE.UU.'),
 ('3 · PROVEEDORES DEL PREDIO — firmar y señar', 'Infraestructura', RURAL, 'Armado de stands',
  9270.00, 'a definir', 'Transferencia', 'Arq. Christian Riccio', '+54 9 11 3481-2179',
  'criccio@larural.com.ar · falta pedirle los espejos.'),
 ('3 · PROVEEDORES DEL PREDIO — firmar y señar', 'Servicios', 'Road Seguridad S.A.', '263 h de vigilancia + planos + supervisión',
  5784.70, 'a definir', 'Transferencia', 'Javier Celiz', 'Cel 11 6303-7845 · Of. 4756-5691',
  'presupuestos@roadseguridad.com.ar · versión V3 del 12/08, lista para firmar.'),
 ('3 · PROVEEDORES DEL PREDIO — firmar y señar', 'Servicios', 'Vittal', 'Servicio médico, 3 días',
  4658.94, 'a definir', 'A resolver', 'vtomas@vittal.com.ar', '—',
  'PENDIENTE: tampoco confirmaron si pueden facturar al exterior.'),
 ('3 · PROVEEDORES DEL PREDIO — firmar y señar', 'Servicios', 'Higia Eventos', 'Limpieza, 205,5 h — ANTICIPO 50%',
  1349.21, 'antes del armado', 'Transferencia', 'Malena Carrizo / Yanina Cruz', 'malena.carrizo@higiaeventos.com',
  'Piden 50% de anticipo antes del armado. El saldo (US$1.349) va contra el evento.'),
 ('3 · PROVEEDORES DEL PREDIO — firmar y señar', 'Servicios', 'Gale Servicios', 'Retiro de residuos — 100% ADELANTADO',
  1179.24, 'antes del evento', 'Transferencia', 'Yanina Cruz', 'info@galeservicios.com',
  'Se factura 100% por adelantado. Trabajan con ARCILLEX; hay que estar dado de alta como generador.'),
 ('3 · PROVEEDORES DEL PREDIO — firmar y señar', 'Servicios', RURAL, 'Conectividad WiFi, 2 redes privadas',
  1174.79, 'a definir', 'Transferencia', 'Gonzalo', 'conectividad@larural.com.ar',
  'Presupuesto N.º 2 del 05/08. Ajusta por IPC.'),
 ('3 · PROVEEDORES DEL PREDIO — firmar y señar', 'Servicios', 'Chanes Seguros', 'Seguro de accidentes, 150 personas',
  631.23, 'antes del armado', 'Transferencia', 'Néstor', 'nestor@chanesseguros.com.ar',
  'Lo exige La Rural para habilitar. Premio único.'),
 ('3 · PROVEEDORES DEL PREDIO — firmar y señar', 'Mobiliario', 'A definir', 'Montaje, acomodación y desmontaje de sillas',
  596.03, 'a definir', 'Transferencia', '—', '—',
  'EN NEGOCIACIÓN: cerrar el número antes de firmar las sillas.'),

 ('4 · PROGRAMADO', 'Merch', 'REMERASYESTAMPADOS', 'Pañuelos — saldo 50% (parte Argentina)',
  3429.73, '13 sep', 'USDT BEP20', 'Leandro Petracci', 'Uruguay 41, Villa Martelli',
  'Factura 2601019 por 5.276,51 USDT; a Argentina le toca el 65%.'),
 ('4 · PROGRAMADO', 'Sede', RURAL, 'Saldo del contrato CC-00957, cuota 2 de 2',
  15777.00, '23 sep', 'Transferencia', 'Victoria Grosi', 'vgrosi@larural.com.ar', ''),
 ('4 · PROGRAMADO', 'Merch', 'DERQUI IMPRESIONES', 'Gráfica Argentina — saldo',
  1192.37, '13 sep', 'USDT BEP20', 'Germán Marino · CUIT 20-37540906-3', 'Charrúa 3366, CABA',
  'A CONFIRMAR: la factura 2601014 viene por el total, no por el saldo.'),
 ('4 · PROGRAMADO', 'Merch', 'LEOTEX', 'Lanyards — saldo',
  816.10, '13 sep', 'USDT BEP20', 'DNI 33914106', 'Marengo 4178, Villa Ballester',
  'A CONFIRMAR: la factura 2601018 viene por el total, no por el saldo.'),

 ('5 · RESTO DEL SALDO — sin fecha acordada', 'Técnica', MET, 'Producción técnica — saldo del 50%',
  30000.00, 'a definir', 'Transferencia / Mercury', 'Lalo Aizenberg', '11 6411-6748',
  'Falta el PDF con el alcance final.'),
 ('5 · RESTO DEL SALDO — sin fecha acordada', 'Técnica', MET, 'Entelado del pabellón',
  16225.17, 'a definir', 'Transferencia', 'Lalo Aizenberg', '11 6411-6748', ''),
 ('5 · RESTO DEL SALDO — sin fecha acordada', 'Técnica', MET, 'Circuito cerrado de cámaras (CCTV)',
  6209.00, 'a definir', 'Transferencia', 'Lalo Aizenberg', '11 6411-6748', 'Con grúa incluida.'),
 ('5 · RESTO DEL SALDO — sin fecha acordada', 'Catering', 'Grupo Ambient · Ángeles', 'Catering — saldo del 50%',
  7284.50, 'a definir', 'Transferencia', 'Ángeles', 'angeles@grupoambient.com.ar', ''),
 ('5 · RESTO DEL SALDO — sin fecha acordada', 'Servicios', 'Higia Eventos', 'Limpieza — saldo del 50%',
  1349.20, 'contra el evento', 'Transferencia', 'Malena Carrizo', 'malena.carrizo@higiaeventos.com', ''),

 ('6 · SIN COTIZAR — pedir presupuesto', 'Servicios', 'A definir', 'Ecobaños',
  0.00, 'pedir', '—', '—', '—', 'Decidido que van ecobaños. Sin cotizar: hoy entra en cero.'),
 ('6 · SIN COTIZAR — pedir presupuesto', 'Merch', 'A definir', 'Marquetería y enmarcado',
  0.00, 'pedir', '—', '—', '—', 'Certificaciones y premios. Sin cotizar.'),
 ('6 · SIN COTIZAR — pedir presupuesto', 'Merch', 'A definir', 'Cheques, escarapelas, diplomas, placas y manillas',
  0.00, 'pedir', '—', '—', '—', 'Lo único del merch que las facturas no cubren.'),
 ('6 · SIN COTIZAR — pedir presupuesto', 'Mobiliario', 'A definir', 'Replanteo de sillas con ingeniero',
  0.00, 'pedir', '—', '—', '—', 'Va aparte de la dirección técnica de La Rural.'),
]

FILAS.append(('7 · AJUSTE DE CONCILIACIÓN', 'Merch', '—',
              'Diferencia de tipo de cambio y saldo de gráfica a confirmar',
              round(FALTA - sum(f[4] for f in FILAS), 2), '—', '—', '—', '—',
              'Las facturas de merch se emitieron a dólar $1.535 y los pagos de junio se hicieron a '
              '$1.475, así que las mitades no suman exacto. El grueso es el saldo de la gráfica de '
              'Argentina: el master dice US$1.839 y la factura 2601014 viene por US$2.385.'))

CONTACTOS = [
 ('La Rural — contrato y eventos', 'Victoria Grosi', 'vgrosi@larural.com.ar', 'congresosyeventos@larural.com.ar'),
 ('La Rural — infraestructura', 'Arq. Christian Riccio', 'criccio@larural.com.ar', '+54 9 11 3481-2179'),
 ('La Rural — conectividad', 'Gonzalo', 'conectividad@larural.com.ar', '—'),
 ('Grupo MET — técnica y entelado', 'Lalo Aizenberg', 'lalo@somosgrupomet.com', '11 6411-6748'),
 ('Grupo Ambient — catering', 'Ángeles', 'angeles@grupoambient.com.ar', '—'),
 ('FDL Eventos — sillas', 'Federico', 'fdleventos@gmail.com', '11 4403-0698 / 11 4040-7012'),
 ('Road Seguridad S.A.', 'Javier Celiz', 'presupuestos@roadseguridad.com.ar', 'Cel 11 6303-7845 · Of. 4756-5691/1435'),
 ('Vittal — servicio médico', '—', 'vtomas@vittal.com.ar', '—'),
 ('Higia Eventos — limpieza', 'Malena Carrizo / Yanina Cruz', 'malena.carrizo@higiaeventos.com', 'yanina.cruz@higiaeventos.com'),
 ('Gale Servicios — residuos', 'Yanina Cruz', 'info@galeservicios.com', '—'),
 ('Chanes Seguros', 'Néstor', 'nestor@chanesseguros.com.ar', '—'),
 ('REMERASYESTAMPADOS — pañuelos, remeras, bordado', 'Leandro Petracci', 'CUIT 20-30367082-4', 'Uruguay 41, Villa Martelli'),
 ('DERQUI IMPRESIONES — gráfica', 'Germán Marino', 'CUIT 20-37540906-3', 'Charrúa 3366, CABA'),
 ('LEOTEX — lanyards', '—', 'DNI 33914106', 'Marengo 4178, Villa Ballester'),
 ('TEXTIL RYU — gorras', '—', 'CUIT 30-71927319-6', 'Larrea 377, Balvanera · YA PAGADO'),
]

# ---------------------------------------------------------------- paleta
F = 'Arial'
NAVY   = '1F3864'
TRAMOS = {   # (color de la banda, relleno de las filas)
 '1': ('C0392B', 'FDEDEC'), '2': ('CA6F1E', 'FDF2E9'), '3': ('1F618D', 'EBF5FB'),
 '4': ('117A65', 'E8F6F3'), '5': ('5D6D7E', 'F2F4F4'), '6': ('7D6608', 'FCF3CF'),
 '7': ('85929E', 'F8F9F9'),
}
USD   = '"US$"#,##0.00;("US$"#,##0.00);-'
USD0  = '"US$"#,##0;("US$"#,##0);-'
BLANCO = Font(name=F, size=10, bold=True, color='FFFFFF')
NEGRO  = Font(name=F, size=10)
BOLD   = Font(name=F, size=10, bold=True)
GRIS   = Font(name=F, size=9, color='566573')
THIN   = Side(style='thin', color='D5D8DC')
BOX    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'CHECKLIST'
ws.sheet_view.showGridLines = False

COLS = [('✔', 5), ('Rubro', 15), ('Proveedor', 27), ('Concepto', 46), ('A pagar', 15),
        ('Cuándo', 18), ('Cómo se paga', 19), ('Contacto', 30), ('Teléfono / mail', 33), ('Notas', 70)]
NC = len(COLS)
for j, (t, an) in enumerate(COLS, 1):
    ws.column_dimensions[L(j)].width = an

# ---------------------------------------------------------------- titulo
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NC)
c = ws.cell(1, 1, 'CHECKLIST DE PAGOS · ARGENTINA')
c.font = Font(name=F, size=18, bold=True, color='FFFFFF')
c.alignment = Alignment(vertical='center', indent=1)
for j in range(1, NC + 1): ws.cell(1, j).fill = PatternFill('solid', fgColor=NAVY)
ws.row_dimensions[1].height = 38

ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=NC)
c = ws.cell(2, 2 - 1, 'Cumbre de los Millonarios Conscientes 2026 · La Rural, Pabellón Azul · '
                      'montaje 2 de octubre, evento 3 y 4 · corte al 2 de septiembre · dólar $1.510')
c.font = Font(name=F, size=10, color='FFFFFF')
c.alignment = Alignment(vertical='center', indent=1)
for j in range(1, NC + 1): ws.cell(2, j).fill = PatternFill('solid', fgColor='2E4B76')
ws.row_dimensions[2].height = 22

# ---------------------------------------------------------------- KPIs
URGENTE = sum(f[4] for f in FILAS if f[0].startswith(('1', '2')))
KPIS = [(1, 3, 'COSTO TOTAL DEL EVENTO', TOTAL, 'EAECEE', '283747'),
        (4, 5, 'YA PAGADO', PAGADO, 'D5F5E3', '196F3D'),
        (6, 7, 'FALTA DESEMBOLSAR', FALTA, 'FCF3CF', '7D6608'),
        (8, 10, 'URGENTE · tramos 1 y 2', URGENTE, 'FADBD8', 'B03A2E')]
for c1, c2, etiqueta, valor, relleno, texto in KPIS:
    ws.merge_cells(start_row=4, start_column=c1, end_row=4, end_column=c2)
    a = ws.cell(4, c1, etiqueta)
    a.font = Font(name=F, size=9, bold=True, color=texto)
    a.alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(start_row=5, start_column=c1, end_row=5, end_column=c2)
    b = ws.cell(5, c1, valor)
    b.font = Font(name=F, size=20, bold=True, color=texto)
    b.number_format = USD0
    b.alignment = Alignment(horizontal='center', vertical='center')
    for r in (4, 5):
        for j in range(c1, c2 + 1):
            ws.cell(r, j).fill = PatternFill('solid', fgColor=relleno)
            ws.cell(r, j).border = BOX
ws.row_dimensions[4].height = 20
ws.row_dimensions[5].height = 34

# ---------------------------------------------------------------- encabezados
HDR = 7
for j, (t, an) in enumerate(COLS, 1):
    c = ws.cell(HDR, j, t)
    c.font = BLANCO
    c.fill = PatternFill('solid', fgColor=NAVY)
    c.alignment = Alignment(horizontal='center' if j in (1, 5) else 'left',
                            vertical='center', wrap_text=True, indent=0 if j in (1, 5) else 1)
    c.border = BOX
ws.row_dimensions[HDR].height = 26
ws.freeze_panes = ws.cell(HDR + 1, 4)

# ---------------------------------------------------------------- filas
dv = DataValidation(type='list', formula1='"☐,✔"', allow_blank=True)
ws.add_data_validation(dv)
r = HDR + 1
tramo_actual, ini_tramo, bandas = None, None, []
primera_fila = r
for tr, rubro, prov, conc, monto, vence, como, cont, tel, nota in FILAS:
    if tr != tramo_actual:
        if tramo_actual is not None:
            bandas.append((fila_banda, ini_tramo, r - 1))
        tramo_actual = tr
        color, _ = TRAMOS[tr[0]]
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC - 1)
        c = ws.cell(r, 1, '   ' + tr)
        c.font = Font(name=F, size=11, bold=True, color='FFFFFF')
        c.alignment = Alignment(vertical='center')
        for j in range(1, NC + 1):
            ws.cell(r, j).fill = PatternFill('solid', fgColor=color)
            ws.cell(r, j).border = BOX
        ws.row_dimensions[r].height = 24
        fila_banda = r
        r += 1
        ini_tramo = r
    relleno = PatternFill('solid', fgColor=TRAMOS[tr[0]][1])
    c = ws.cell(r, 1, '☐'); c.font = Font(name=F, size=12); dv.add(c)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(r, 2, rubro).font = GRIS
    ws.cell(r, 3, prov).font = BOLD
    ws.cell(r, 4, conc).font = NEGRO
    c = ws.cell(r, 5, monto if monto else None); c.font = BOLD; c.number_format = USD
    ws.cell(r, 6, vence).font = NEGRO
    ws.cell(r, 7, como).font = GRIS
    ws.cell(r, 8, cont).font = NEGRO
    ws.cell(r, 9, tel).font = NEGRO
    ws.cell(r, 10, nota).font = GRIS
    for j in range(1, NC + 1):
        ws.cell(r, j).fill = relleno
        ws.cell(r, j).border = BOX
        ws.cell(r, j).alignment = Alignment(wrap_text=(j in (4, 10)), vertical='center',
                                            horizontal='center' if j == 1 else 'left',
                                            indent=0 if j == 1 else 1)
    ws.row_dimensions[r].height = 30
    r += 1
bandas.append((fila_banda, ini_tramo, r - 1))
ultima_fila = r - 1

# subtotal de cada tramo, en la propia banda
for fila_banda, ini, fin in bandas:
    c = ws.cell(fila_banda, NC, f'=SUM(E{ini}:E{fin})')
    c.font = Font(name=F, size=12, bold=True, color='FFFFFF')
    c.number_format = USD0
    c.alignment = Alignment(horizontal='right', vertical='center', indent=1)

# ---------------------------------------------------------------- total
r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
c = ws.cell(r, 1, '   TOTAL A DESEMBOLSAR')
c.font = Font(name=F, size=13, bold=True, color='FFFFFF')
c.alignment = Alignment(vertical='center')
c = ws.cell(r, 5, '=' + '+'.join(f'{L(NC)}{b[0]}' for b in bandas))
c.font = Font(name=F, size=13, bold=True, color='FFFFFF')
c.number_format = USD0
for j in range(1, NC + 1):
    ws.cell(r, j).fill = PatternFill('solid', fgColor=NAVY)
    ws.cell(r, j).border = BOX
ws.row_dimensions[r].height = 30
fila_total = r

# tildado: la fila se pinta de verde y se agrisa
verde = DifferentialStyle(fill=PatternFill(bgColor='D5F5E3'),
                          font=Font(color='7B8A8B', italic=True, strike=True))
ws.conditional_formatting.add(
    f'A{primera_fila}:{L(NC)}{ultima_fila}',
    Rule(type='expression', formula=[f'$A{primera_fila}="✔"'], dxf=verde, stopIfTrue=False))

r += 2
ws.cell(r, 2, 'Marcá la columna ✔ a medida que salen los pagos: la fila se pinta de verde sola.').font = GRIS
ws.cell(r + 1, 2, 'Los tramos están ordenados por urgencia. El 1 es lo vencido o ya facturado; el 2, los anticipos del 50% que destraban técnica y catering.').font = GRIS
ws.cell(r + 2, 2, 'Los pagos de merch van en USDT por red BEP20. Remerasyestampados: 0xae874c3db52ca1e45604549a5527f450a877b97a · Leotex y Derqui: 0x7d0a6c347008305af1d6643c2382e1495af7dc0e. Verificá la dirección con el proveedor antes de operar.').font = GRIS
ws.cell(r + 3, 2, 'FDL Eventos y Vittal todavía no confirmaron si pueden facturar a la empresa de EE.UU.: son US$25.333 trabados por un tema administrativo.').font = GRIS
ws.cell(r + 4, 2, 'Higia pide 50% de anticipo antes del armado y Gale factura el 100% por adelantado.').font = GRIS
ws.auto_filter.ref = f'A{HDR}:{L(NC)}{ultima_fila}'

# ================================================================ CONTACTOS
ws2 = wb.create_sheet('CONTACTOS')
ws2.sheet_view.showGridLines = False
CC = [('Proveedor', 46), ('Persona', 30), ('Mail', 38), ('Teléfono / dato fiscal', 40)]
for j, (t, an) in enumerate(CC, 1): ws2.column_dimensions[L(j)].width = an
ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
c = ws2.cell(1, 1, 'CONTACTOS DE PROVEEDORES')
c.font = Font(name=F, size=16, bold=True, color='FFFFFF')
c.alignment = Alignment(vertical='center', indent=1)
for j in range(1, 5): ws2.cell(1, j).fill = PatternFill('solid', fgColor=NAVY)
ws2.row_dimensions[1].height = 34
for j, (t, an) in enumerate(CC, 1):
    c = ws2.cell(3, j, t); c.font = BLANCO; c.fill = PatternFill('solid', fgColor=NAVY)
    c.alignment = Alignment(vertical='center', indent=1); c.border = BOX
ws2.row_dimensions[3].height = 22
r2 = 4
for prov, per, mail, tel in CONTACTOS:
    ws2.cell(r2, 1, prov).font = BOLD
    ws2.cell(r2, 2, per).font = NEGRO
    ws2.cell(r2, 3, mail).font = NEGRO
    ws2.cell(r2, 4, tel).font = NEGRO
    relleno = PatternFill('solid', fgColor='FFFFFF' if r2 % 2 else 'F4F6F7')
    for j in range(1, 5):
        ws2.cell(r2, j).fill = relleno; ws2.cell(r2, j).border = BOX
        ws2.cell(r2, j).alignment = Alignment(vertical='center', indent=1)
    ws2.row_dimensions[r2].height = 22
    r2 += 1
ws2.freeze_panes = ws2.cell(4, 1)

# ================================================================ RESUMEN
ws3 = wb.create_sheet('RESUMEN')
ws3.sheet_view.showGridLines = False
for j, an in enumerate([42, 18, 30, 12], 1): ws3.column_dimensions[L(j)].width = an
ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
c = ws3.cell(1, 1, 'RESUMEN POR TRAMO')
c.font = Font(name=F, size=16, bold=True, color='FFFFFF')
c.alignment = Alignment(vertical='center', indent=1)
for j in range(1, 5): ws3.cell(1, j).fill = PatternFill('solid', fgColor=NAVY)
ws3.row_dimensions[1].height = 34
for j, t in enumerate(['Tramo', 'USD', '', '% del total'], 1):
    c = ws3.cell(3, j, t); c.font = BLANCO; c.fill = PatternFill('solid', fgColor=NAVY)
    c.alignment = Alignment(vertical='center', indent=1); c.border = BOX
r3 = 4
ini_r3 = r3
for (fila_banda, ini, fin), (tr, *_) in zip(bandas, [(f[0],) for f in FILAS if True][:0] or
                                            [(b,) for b in dict.fromkeys(f[0] for f in FILAS)]):
    color = TRAMOS[tr[0]][0]
    ws3.cell(r3, 1, tr).font = BOLD
    c = ws3.cell(r3, 2, f"=CHECKLIST!{L(NC)}{fila_banda}"); c.font = BOLD; c.number_format = USD0
    c = ws3.cell(r3, 3, f'=REPT("|",ROUND(B{r3}/$B${ini_r3 + len(bandas) + 1}*60,0))')
    c.font = Font(name=F, size=11, color=color)
    c = ws3.cell(r3, 4, f'=B{r3}/$B${ini_r3 + len(bandas) + 1}'); c.font = NEGRO; c.number_format = '0.0%'
    for j in range(1, 5):
        ws3.cell(r3, j).fill = PatternFill('solid', fgColor=TRAMOS[tr[0]][1])
        ws3.cell(r3, j).border = BOX
        ws3.cell(r3, j).alignment = Alignment(vertical='center', indent=1)
    ws3.row_dimensions[r3].height = 22
    r3 += 1
r3 += 1
ws3.cell(r3, 1, 'TOTAL A DESEMBOLSAR').font = Font(name=F, size=12, bold=True, color='FFFFFF')
c = ws3.cell(r3, 2, f'=SUM(B{ini_r3}:B{r3 - 2})')
c.font = Font(name=F, size=12, bold=True, color='FFFFFF'); c.number_format = USD0
for j in range(1, 5):
    ws3.cell(r3, j).fill = PatternFill('solid', fgColor=NAVY); ws3.cell(r3, j).border = BOX
    ws3.cell(r3, j).alignment = Alignment(vertical='center', indent=1)
ws3.row_dimensions[r3].height = 28

wb.save('checklist_pagos_argentina.xlsx')

# ================================================================ CSV
filas = [['CHECKLIST DE PAGOS · ARGENTINA · CUMBRE DE LOS MILLONARIOS CONSCIENTES 2026'],
         ['La Rural, Pabellón Azul · corte al 2 de septiembre · dólar $1.510'], [],
         ['Costo total del evento', round(TOTAL, 2)], ['Ya pagado', round(PAGADO, 2)],
         ['FALTA DESEMBOLSAR', round(FALTA, 2)], [],
         ['Pagado'] + [c[0] for c in COLS[1:]] + ['Tramo']]
for tr, rubro, prov, conc, monto, vence, como, cont, tel, nota in FILAS:
    filas.append(['☐', rubro, prov, conc, round(monto, 2), vence, como, cont, tel, nota, tr])
filas += [[], ['TOTAL A DESEMBOLSAR', '', '', '', round(sum(f[4] for f in FILAS), 2)], [],
          ['CONTACTOS'], ['Proveedor', 'Persona', 'Mail', 'Teléfono / dato fiscal']]
for x in CONTACTOS: filas.append(list(x))
sal = io.StringIO(); csv.writer(sal, lineterminator='\n').writerows(filas)
io.open('checklist_pagos_argentina.csv', 'w', encoding='utf-8').write(sal.getvalue())

print('Costo total      {:>12,.2f}'.format(TOTAL))
print('Ya pagado        {:>12,.2f}'.format(PAGADO))
print('Falta            {:>12,.2f}'.format(FALTA))
print('Suma del detalle {:>12,.2f}'.format(sum(f[4] for f in FILAS)))
print('Urgente (1+2)    {:>12,.2f}'.format(URGENTE))
print('filas de detalle:', len(FILAS), '· tramos:', len(bandas), '· ultima fila:', ultima_fila)
for (fb, ini, fin), tr in zip(bandas, dict.fromkeys(f[0] for f in FILAS)):
    print('  banda fila {:>3}  SUM(E{}:E{})  {:<44} {:>12,.2f}'.format(
        fb, ini, fin, tr, sum(f[4] for f in FILAS if f[0] == tr)))
