# -*- coding: utf-8 -*-
"""
Consolida TODOS los costos del evento de Argentina (La Rural, 3-4 oct 2026) en un solo libro.

Fuentes:
  A. Cotizaciones reales recibidas por mail (agustinalorenzog@gmail.com), jun-ago 2026.
  B. Hoja "Argentina (3 oct-4 oct)" de INVERSIÓN GIRA CUMBRE 2026.xlsx (estimados en USD).
  C. Hoja PAGOS del mismo archivo (contratos con anticipo ya pagado).
"""
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as L
from openpyxl.worksheet.datavalidation import DataValidation

HOY = datetime.date(2026, 9, 2)
TC_DEFECTO = 1510.0   # dolar operativo con el que trabaja Agustina (verificado: $900.000 = US$596 y $24.500.000 = US$16.225)

F = 'Arial'
AZUL  = Font(name=F, size=10, color='0000FF')
NEGRO = Font(name=F, size=10)
VERDE = Font(name=F, size=10, color='008000')
BOLD  = Font(name=F, size=10, bold=True)
SUB   = Font(name=F, size=9, italic=True, color='595959')
TIT   = Font(name=F, size=14, bold=True, color='1F3864')
HDR   = Font(name=F, size=10, bold=True, color='FFFFFF')
FH    = PatternFill('solid', fgColor='1F3864')
FPAIS = PatternFill('solid', fgColor='1F3864')
FBLOQ = PatternFill('solid', fgColor='DCE6F1')
FGRIS = PatternFill('solid', fgColor='F2F2F2')
FAMAR = PatternFill('solid', fgColor='FFEB9C')
FROJO = PatternFill('solid', fgColor='FFC7CE')
FVERD = PatternFill('solid', fgColor='C6EFCE')
FIN   = PatternFill('solid', fgColor='FFFF00')
THIN  = Side(style='thin', color='BFBFBF')
BOX   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
USD   = '"US$"#,##0.00;("US$"#,##0.00);-'
USD0  = '"US$"#,##0;("US$"#,##0);-'
ARS   = '"$"#,##0;("$"#,##0);-'
PCT   = '0.0%'
FECHA = 'DD/MM/YYYY'

wb = openpyxl.Workbook()

def hoja(nombre, titulo, bajada):
    ws = wb.create_sheet(nombre)
    ws['A1'] = titulo; ws['A1'].font = TIT
    ws['A2'] = bajada; ws['A2'].font = SUB
    return ws

def encabezados(ws, fila, cols):
    for j, (t, an) in enumerate(cols, 1):
        c = ws.cell(fila, j, t); c.font = HDR; c.fill = FH; c.border = BOX
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.column_dimensions[L(j)].width = an
    ws.freeze_panes = ws.cell(fila + 1, 1)

def banda(ws, fila, texto, ncols):
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=ncols)
    c = ws.cell(fila, 1, texto)
    c.font = Font(name=F, size=11, bold=True, color='FFFFFF'); c.fill = FPAIS
    c.alignment = Alignment(vertical='center', indent=1)
    ws.row_dimensions[fila].height = 20
    for j in range(1, ncols + 1): ws.cell(fila, j).fill = FPAIS
    return fila + 1

def banda_bloque(ws, fila, texto, ncols):
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=ncols)
    c = ws.cell(fila, 1, f'  {texto}')
    c.font = Font(name=F, size=9, bold=True, color='1F3864'); c.fill = FBLOQ
    for j in range(1, ncols + 1): ws.cell(fila, j).fill = FBLOQ
    return fila + 1

# ============================================================================
# DATOS
# ============================================================================
# (bloque, rubro, proveedor, detalle, moneda, monto_sin_iva, iva, estado,
#  estimado_usd_sheet, fuente)
# estado: Contratado / Cotizado / Alternativa / Sin cotizar
ARS_, USD_ = 'ARS', 'USD'
# estado: Cerrado / Contratado / Cotizado / En revisión / En negociación / Bonificado / Alternativa / Sin cotizar
ITEMS = [
 # ---------------- SEDE ----------------
 ('Sede', 'Alquiler Pabellón Azul', 'La Rural (contrato CC-00957)',
  'Locación 2 días de evento + 1 día de montaje', USD_, 75000.0, 0.0, 'Contratado',
  70000.0, 'Contrato CC-00957 · hoja PAGOS del master (US$43.446 abonados el 27/04)'),
 ('Infraestructura', 'Armado de stands', 'La Rural — Jefatura de Infraestructura',
  'Armado del stand de staff (100 m²), el de producción (menos de 50 m²) y el mobiliario del camarín de speakers. Falta pedirles los espejos. Las sillas del público van en su propia línea.',
  USD_, 9270.0, 0.0, 'Cerrado', 0.0,
  'Definición de Agustina sobre el presupuesto de $66.494.708 del 17/08 · ver hoja AHORROS'),
 ('Mobiliario', 'Sillas del público (5.500)', 'FDL Eventos',
  'Presupuesto 026-2697, alquiler por 4 días: 1.000 sillas hotel caño gris tapizado negro ($7.000.000) + 4.500 sillas plásticas Munro reforzadas ($17.100.000) + flete ($1.700.000). NO incluye armado.',
  ARS_, 25800000.0, 0.21, 'Cotizado', 16500.0,
  'Presupuesto FDL Eventos N.º 026-2697 del 07/08/2026 · pendiente de resolver facturación al exterior'),
 ('Mobiliario', 'Montaje, acomodación y desmontaje de sillas', 'En negociación',
  'Montaje previo, acomodación el sábado a la noche después de la dinámica y desmontaje el domingo.',
  ARS_, 900000.0, 0.0, 'En negociación', 0.0,
  'Negociación de Agustina · cerrar el número antes de firmar las sillas'),
 ('Mobiliario', 'Replanteo de sillas (ingeniero)', 'Sin cotizar',
  'Replanteo del layout de sillas del pabellón. Confirmado que va aparte de la dirección técnica de La Rural.',
  USD_, 0.0, 0.0, 'Sin cotizar', 0.0, 'Pendiente de pedir presupuesto'),
 ('Servicios', 'Conectividad WiFi', 'La Rural — Servicios de Conectividad',
  '2 redes privadas de 30 Mbps (técnica + staff/acreditación), del 2 al 4 de octubre',
  ARS_, 1466063.84, 0.21, 'Cotizado', 800.0,
  'Mail conectividad@larural.com.ar 05/08/2026 · presupuesto N.º 2, base IPC junio (el del 16/06 era $1.438.728)'),
 # ---------------- TÉCNICA ----------------
 ('Técnica', 'Producción técnica', 'Cerrado por Agustina',
  'Sonido, iluminación, video y LED. Cerrado en dólares.',
  USD_, 60000.0, 0.0, 'Cerrado', 34000.0,
  'Negociación de Agustina · reemplaza la cotización vencida de Prina'),
 ('Técnica', 'Entelado', 'Cerrado por Agustina',
  'Entelado del pabellón. Negociado a la mitad de lo que estaba cotizado.',
  ARS_, 24500000.0, 0.0, 'Cerrado', 0.0,
  'Negociación de Agustina · US$16.000 aprox. al dólar de referencia'),
 ('Técnica', 'Circuito cerrado (CCTV)', 'Cerrado por Agustina',
  'Circuito cerrado de cámaras para el pabellón.',
  USD_, 6209.0, 0.0, 'Cerrado', 0.0, 'Negociación de Agustina'),
 ('Técnica', 'Bonificado por el proveedor de técnica', 'Sin costo',
  '500 vallas y efectos especiales, sin cargo. Es lo que evita el gasto de vallado que estaba presupuestado.',
  USD_, 0.0, 0.0, 'Bonificado', 0.0, 'Negociación de Agustina'),
 # ---------------- SERVICIOS ----------------
 ('Servicios', 'Seguridad y vigilancia', 'Road Seguridad S.A.',
  '263 horas de vigilancia (armado, evento y desarme) + planos de evacuación $715.000 + supervisión general $795.000',
  ARS_, 7218928.96, 0.21, 'Cotizado', 0.0,
  'Mail presupuestos@roadseguridad.com.ar 12/08/2026 (versión V3, lista para firmar)'),
 ('Servicios', 'Servicio médico', 'Vittal — Socorro Médico Privado',
  '1 UTIM TRI con dos médicos + enfermero/chofer, los tres días (2, 3 y 4 de octubre)',
  ARS_, 6366521.27, 0.105, 'Cotizado', 0.0,
  'Mail vtomas@vittal.com.ar 28/08/2026 · tarifas actualizadas, sin consultorio'),
 ('Servicios', 'Limpieza', 'Higia Eventos',
  '205,5 horas de limpieza: auditorio, baños y guardarropas, del 2 al 5 de octubre',
  ARS_, 3367440.91, 0.21, 'Cotizado', 0.0,
  'Mail higiaeventos 22/06/2026 (reenviado 29/06) · 50% de anticipo antes del armado'),
 ('Servicios', 'Retiro de residuos', 'Gale Servicios',
  'Roll off con disposición final incluida. Cotización base abril 2026.',
  ARS_, 1471608.62, 0.21, 'Cotizado', 0.0,
  'Mail higiaeventos 22/06/2026 (adjunto Gale) · se factura 100% por adelantado'),
 ('Servicios', 'Baños químicos — ecobaños', 'Sin cotizar',
  'Decidido: ecobaños, no los químicos tradicionales. Más limpios, mejor a la vista y ecofriendly. Falta pedir presupuesto.',
  USD_, 0.0, 0.0, 'Sin cotizar', 0.0, 'Decisión de Agustina · pendiente de cotizar'),
 ('Servicios', 'Seguro de accidentes personales', 'Chanes Seguros',
  'Cobertura MIA (muerte, invalidez y asistencia médica) para 150 personas, exigida por La Rural',
  ARS_, 953151.93, 0.0, 'Cotizado', 400.0,
  'Mail nestor@chanesseguros.com.ar 13/07/2026 · premio único, cotizado para 15-19/07'),
 # ---------------- CATERING ----------------
 ('Catering', 'Catering VIP, staff y equipo', 'Grupo Ambient',
  'Número final por las dos propuestas juntas: 600 lunchbox para los VIP + desayuno, almuerzo y cena para 150 personas.',
  USD_, 14569.0, 0.0, 'Cerrado', 33820.0,
  'Cierre de Agustina con Grupo Ambient · reemplaza la estimación de $25.000.000'),
 ('Catering', 'Food trucks', 'Sin costo',
  'No los paga la producción.',
  USD_, 0.0, 0.0, 'Bonificado', 0.0, 'Confirmado por Agustina'),
 # ---------------- MERCH ----------------
 # Todo el merch sale de la carpeta de Drive "Presupuestos merch" (facturas 2601009 a 2601019).
 # Los proveedores son argentinos y facturan a IIDAI LLC; el master los tenia como "Colombia".
 ('Merch', 'Pañuelos / pañoletas estampados (8.500)', 'REMERASYESTAMPADOS',
  'Factura 2601012 del 20/06. Compra compartida Argentina + Uruguay: se imputa a Argentina el 65% (5.500 de 8.500 unidades, la misma proporcion que la grafica). El 50% ya esta abonado; el saldo se paga con la factura 2601019 por 5.276,51 USDT, de los cuales 3.429,73 son de Argentina.',
  ARS_, 13387500.0, 0.21, 'Contratado', 0.0,
  'Drive · Presupuesto Pañuelos - Arg y Uru · facturas 2601012 (20/06) y 2601019 (16/08)', 0.65),
 ('Merch', 'Remeras premium estampadas (250)', 'REMERASYESTAMPADOS',
  'Factura 2601017 del 16/08, $14.110 por unidad. Sube de 230 a 250 unidades contra la factura original 2601011, al mismo precio unitario. SIN PAGAR: se debe el 100% (2.780,64 USDT).',
  ARS_, 3527500.0, 0.21, 'Contratado', 0.0,
  'Drive · Presupuestos camisetas + pañoletas · factura 2601017 (16/08)', 1.0),
 ('Merch', 'Bordado de gorras (940)', 'REMERASYESTAMPADOS',
  'Bordado computarizado sobre las gorras que provee el cliente. Presupuesto del 03/07 con el descuento de $411.800 ya aplicado. Precio final, no lleva IVA. SIN PAGAR: se debe el 100% (1.721,95 USDT).',
  ARS_, 2643200.0, 0.0, 'Contratado', 0.0,
  'Drive · Presupuesto gorras (compra y bordado) / Bordado · factura 2601016 (10/08)', 1.0),
 ('Merch', 'Gorras Flex unicolor/bicolor (940)', 'TEXTIL RYU S.R.L.',
  'Factura 2601009 del 20/06, pagada al 100% el 25/06. Es la compra de las gorras; el bordado va en su propia linea.',
  ARS_, 2224980.0, 0.21, 'Contratado', 5000.0,
  'Drive · Compra de gorras · factura 2601009 · hoja PAGOS del master', 1.0),
 ('Merch', 'Lanyards premium doble raso 25 mm (1.360)', 'LEOTEX',
  'Sublimacion full print en ambas caras, 1.300 negros + 50 azules (la factura dice 1.360, la suma da 1.350: hay 10 de diferencia). Reemitida como 2601018 en USDT. OJO: la carpeta se llama "Nuevo pago - 50%" pero la factura viene por el total.',
  ARS_, 2070600.0, 0.21, 'Contratado', 0.0,
  'Drive · Presupuesto Lanyards · facturas 2601013 (20/06) y 2601018 (16/08)', 1.0),
 ('Merch', 'Grafica oficial CMC — Argentina', 'DERQUI IMPRESIONES',
  '5.500 hojas A4, 5.500 tarjetas, 1.030 credenciales, 5.500 mapas y 1.000 bolsas de friselina. Reemitida el 16/08 para cambiar el medio de pago a USDT.',
  ARS_, 3025260.0, 0.21, 'Contratado', 0.0,
  'Drive · Factura/presu ARGENTINA · factura 2601014 (20/06, reemitida 16/08)', 1.0),
 ('Merch', 'Grafica oficial CMC — Uruguay', 'DERQUI IMPRESIONES',
  'NO ES DE ARGENTINA. 2.800 hojas A4, 2.800 tarjetas, 530 credenciales, 2.800 mapas y 1.000 bolsas de friselina, para el evento de Uruguay. Se deja a la vista en cero porque en la hoja PAGOS del master estaba cargada como Argentina: es el monto que no cerraba.',
  ARS_, 1564626.0, 0.21, 'Otro pais', 0.0,
  'Drive · Factura/presu URUGUAY · factura 2601015 (20/06, reemitida 16/08)', 0.0),
 ('Merch', 'Cintas portacredencial (1.300) — alternativa', 'Blocko',
  'Cinta raso 25 mm impresa a 4 colores dos caras + mosqueton zamak, $1.346 c/u. Es el MISMO item que los lanyards de LEOTEX, que ya esta contratado y con anticipo pagado. Estaba sumando dos veces: ahora queda como alternativa descartada.',
  ARS_, 1749800.0, 0.21, 'Alternativa', 0.0,
  'Mail hola@blocko.com.ar 09/06/2026 · descartada contra la factura 2601013 de LEOTEX', 1.0),
 ('Merch', 'Marqueteria y enmarcado', 'Sin cotizar',
  'Enmarcado de las certificaciones y los premios. Rubro nuevo, no estaba en ninguna planilla.',
  USD_, 0.0, 0.0, 'Sin cotizar', 0.0, 'Agregado por Agustina · pendiente de cotizar', 1.0),
 ('Merch', 'Cheques, escarapelas, diplomas, placas y manillas', 'Sin cotizar',
  'Lo unico del merch que las facturas de Drive NO cubren. Las camisetas, los mapas, las bolsas y las credenciales ya estan facturados arriba.',
  USD_, 0.0, 0.0, 'Sin cotizar', 0.0,
  'Pendiente de cotizar · antes estaba dentro del estimado global de US$8.689', 1.0),
 ('Merch', 'Estimado global de merch del master (reemplazado)', 'Varios',
  'La hoja Argentina del master tenia US$8.689 en un solo renglon por contratos, cheques, escarapelas, bolsas, camisetas, mapas, diplomas, placas y manillas. Queda en cero como costo porque las facturas reales de arriba lo reemplazan; se conserva el estimado para poder comparar contra el presupuesto original.',
  USD_, 0.0, 0.0, 'Reemplazado', 8689.0,
  'Hoja Argentina del master · reemplazado por las facturas 2601009 a 2601019', 1.0),
 # ---------------- PRODUCCIÓN ----------------
 ('Producción', 'Vallado (500 vallas)', 'Bonificado por técnica',
  'Las 500 vallas las regala el proveedor de técnica. Estaban presupuestadas en US$1.250.',
  USD_, 0.0, 0.0, 'Bonificado', 1250.0, 'Negociación de Agustina'),
 ('Producción', 'Intercoms', 'Proveedor a confirmar',
  'Ya comprados y pagados al 100%. El master los tenía en 2 cuotas al 24/08 y 23/09, sin abonar.',
  USD_, 560.0, 0.0, 'Cerrado', 0.0, 'Confirmado por Agustina el 03/09'),
]

# ============================================================================
# 1. LEEME
# ============================================================================
ws = hoja('LEEME', 'Costo total del evento de Argentina · CMC 2026',
          'La Rural, Pabellón Azul — montaje 2 de octubre, evento 3 y 4 de octubre de 2026.')
ws.column_dimensions['A'].width = 30; ws.column_dimensions['B'].width = 115
filas = [
 ('Qué es esto', 'Unifica los presupuestos que llegaron por mail (jun-ago 2026) con la hoja "Argentina (3 oct-4 oct)" del master de inversión, para poder decir cuánto sale el evento de verdad y cuánto por persona.'),
 ('Fecha de corte', 'Todo calculado al 02/09/2026.'),
 ('Tipo de cambio', 'Celda TABLERO!C4, editable. $1.510, que es el dólar con el que viene trabajando Agustina. Se verificó contra dos de sus cifras: el montaje de $900.000 le da US$596 y el entelado de $24.500.000 le da US$16.225. Todas las conversiones de pesos a dólares salen de esa única celda.'),
 ('', ''),
 ('Cómo leer los estados', 'Contratado = hay contrato o anticipo pagado. Cotizado = hay presupuesto formal por mail. Alternativa = segunda opción para el mismo ítem, NO suma al total. Otro país = es un gasto de otro evento de la gira, entra en cero. Reemplazado = estimado viejo que ya está cubierto por facturas reales, entra en cero. Sin cotizar = sólo hay un estimado del sheet, o ni eso.'),
 ('Columna "% Argentina"', 'Es la parte de cada compra que se le imputa a este evento. Vale 100% en casi todo. Los pañuelos son una compra compartida con Uruguay y van al 65% (5.500 de 8.500 unidades); la gráfica de Uruguay va al 0%.'),
 ('Regla del total', 'El total suma los ítems Contratado + Cotizado + Sin cotizar. Los marcados "Alternativa" quedan fuera para no duplicar (las sillas de La Rural vs las de FDL, y las LED de VMG contra la técnica integral de Prina).'),
 ('IVA', 'Las cotizaciones argentinas casi todas vienen sin IVA. La columna IVA aplica la alícuota de cada una (21% general, 10,5% el servicio médico) para llegar al costo real.'),
 ('', ''),
 ('Advertencia 1 — vencimientos', 'La cotización de Prina (técnica, el ítem más caro en pesos) tenía validez de 10 días y venció en junio. VMG se reserva recotizar si el dólar salta más de 15%. La Rural ajusta el saldo por IPC. Los pesos de este libro son de la fecha de cada cotización, no de hoy.'),
 ('Advertencia 2 — huecos', 'Esta hoja tiene sólo lo que cargó Agustina. Los ecobaños, la marquetería, el replanteo de sillas y los cheques/escarapelas/diplomas/placas/manillas están definidos pero todavía sin precio, así que entran en cero. El costo por persona es, por lo tanto, un piso.'),
 ('', ''),
 ('Merch', 'Todo el merch sale de la carpeta de Drive "Presupuestos merch": nueve facturas reales (2601009 a 2601019) que reemplazan el renglón único de US$8.689 que traía el master. Los proveedores son argentinos (Remerasyestampados, Textil Ryu, Leotex, Derqui), no colombianos como decía la hoja PAGOS. Ver la hoja MERCH.'),
 ('Merch — qué se pagó', 'Las gorras están pagadas al 100%. Pañuelos, lanyards y gráfica de Argentina tienen el 50% abonado. Las remeras y el bordado de las gorras no tienen nada pagado: se debe el 100%.'),
 ('Merch — dos ajustes', 'Se sacaron dos duplicaciones: las cintas de Blocko eran el mismo ítem que los lanyards de Leotex ya contratados, y la gráfica de Uruguay (US$1.233) estaba cargada en el master con país = Argentina.'),
 ('Advertencia 3 — facturación', 'FDL Eventos (sillas) y Vittal (servicio médico) todavía no confirmaron si pueden facturar a la empresa de EE.UU. Si no pueden, hay que resolver el circuito de pago antes de cerrar.'),
 ('', ''),
 ('Fuentes de mail', 'La Rural infraestructura (criccio@larural.com.ar, 17/08) · La Rural conectividad (16/06) · Prina (18/06) · VMG (29/06) · Road Seguridad (12/08) · Vittal (28/08) · Higia + Gale (22/06) · Chanes Seguros (13/07) · Blocko (09/06) · FDL Eventos (07/08) · AmbientHouse (16/06).'),
 ('Fuente de planilla', 'INVERSIÓN GIRA CUMBRE 2026.xlsx — hojas "Argentina (3 oct-4 oct)" y "PAGOS".'),
]
r = 4
for a, b in filas:
    ws.cell(r, 1, a).font = BOLD if a else NEGRO
    c = ws.cell(r, 2, b); c.font = NEGRO; c.alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[r].height = 30 if b else 8
    r += 1

# ============================================================================
# 2. COSTOS UNIFICADOS
# ============================================================================
ws = hoja('COSTOS UNIFICADOS', 'Todos los costos de Argentina, bloque por bloque',
          'Una fila por concepto. "Monto a usar" toma la cotización real cuando existe y el estimado del sheet cuando no.')
COLS = [('Rubro', 34), ('Proveedor', 30), ('Detalle', 62), ('Estado', 13), ('Moneda', 8),
        ('Monto sin IVA', 15), ('IVA', 8), ('Total moneda origen', 17), ('% Argentina', 11),
        ('Total USD', 14), ('Estimado del sheet\n(USD)', 15), ('Diferencia', 14), ('Fuente', 62)]
NC = len(COLS)
encabezados(ws, 4, COLS)

orden_bloques = ['Sede', 'Infraestructura', 'Mobiliario', 'Técnica', 'Servicios', 'Merch', 'Catering', 'Producción', 'Equipo']
r = 5
filas_detalle, subtotales = [], []
for bloque in orden_bloques:
    grupo = [x for x in ITEMS if x[0] == bloque]
    if not grupo: continue
    r = banda(ws, r, bloque.upper(), NC)
    ini = r
    for it in grupo:
        (_, rubro, prov, det, mon, monto, iva, estado, est, fuente) = it[:10]
        pct = it[10] if len(it) > 10 else 1.0
        ws.cell(r, 1, rubro).font = BOLD
        ws.cell(r, 2, prov).font = AZUL
        ws.cell(r, 3, det).font = NEGRO
        ws.cell(r, 4, estado).font = BOLD
        ws.cell(r, 5, mon).font = AZUL
        c = ws.cell(r, 6, monto or None); c.font = AZUL; c.number_format = ARS if mon == ARS_ else USD
        c = ws.cell(r, 7, iva or None); c.font = AZUL; c.number_format = PCT
        c = ws.cell(r, 8, f'=IF(F{r}="","",F{r}*(1+N(G{r})))'); c.font = NEGRO
        c.number_format = ARS if mon == ARS_ else USD
        c = ws.cell(r, 9, pct); c.font = AZUL if pct != 1.0 else SUB; c.number_format = PCT
        c = ws.cell(r, 10, f'=IF(OR(D{r}="Bonificado",D{r}="Reemplazado"),0,'
                           f'N(I{r})*IF(H{r}="",N(K{r}),IF(E{r}="USD",H{r},H{r}/TABLERO!$C$4)))')
        c.font = NEGRO; c.number_format = USD
        c = ws.cell(r, 11, est or None); c.font = AZUL; c.number_format = USD
        c = ws.cell(r, 12, f'=IF(D{r}="Alternativa","",J{r}-N(K{r}))'); c.font = NEGRO; c.number_format = USD
        ws.cell(r, 13, fuente).font = SUB
        relleno = {'Cerrado': FVERD, 'Contratado': FVERD, 'Cotizado': None, 'En revisión': FAMAR,
                   'En negociación': FAMAR, 'Bonificado': FGRIS, 'Alternativa': FGRIS,
                   'Otro pais': FGRIS, 'Reemplazado': FGRIS, 'Sin cotizar': FAMAR}[estado]
        if relleno:
            for j in range(1, NC + 1): ws.cell(r, j).fill = relleno
        for j in range(1, NC + 1):
            ws.cell(r, j).border = BOX
            ws.cell(r, j).alignment = Alignment(wrap_text=True, vertical='top')
        ws.row_dimensions[r].height = 42
        filas_detalle.append((r, estado))
        r += 1
    ws.cell(r, 1, f'Subtotal {bloque}').font = BOLD
    c = ws.cell(r, 10, f'=SUMIF($D${ini}:$D${r-1},"<>Alternativa",$J${ini}:$J${r-1})')
    c.font = BOLD; c.number_format = USD
    c = ws.cell(r, 11, f'=SUM($K${ini}:$K${r-1})'); c.font = BOLD; c.number_format = USD
    c = ws.cell(r, 12, f'=J{r}-K{r}'); c.font = BOLD; c.number_format = USD
    for j in range(1, NC + 1): ws.cell(r, j).fill = FGRIS; ws.cell(r, j).border = BOX
    subtotales.append(r)
    r += 1

r += 1
ws.cell(r, 1, 'COSTO TOTAL DEL EVENTO').font = Font(name=F, size=12, bold=True)
for col, letra in ((10, 'J'), (11, 'K'), (12, 'L')):
    c = ws.cell(r, col, '=' + '+'.join(f'{letra}{x}' for x in subtotales))
    c.font = Font(name=F, size=12, bold=True); c.number_format = USD0; c.fill = FVERD; c.border = BOX
fila_total = r
ws.auto_filter.ref = f'A4:{L(NC)}4'

# ============================================================================
# 3. TABLERO
# ============================================================================
ws = hoja('TABLERO', 'Argentina en una pantalla', 'La Rural · Pabellón Azul · 3 y 4 de octubre de 2026')
ws.column_dimensions['A'].width = 42
for col in 'BCD': ws.column_dimensions[col].width = 18
ws.column_dimensions['E'].width = 76

ws['A4'] = 'Tipo de cambio (ARS por USD)'; ws['A4'].font = BOLD
ws['C4'] = TC_DEFECTO; ws['C4'].font = AZUL; ws['C4'].fill = FIN; ws['C4'].number_format = '#,##0.00'
ws['E4'] = 'Dólar con el que trabaja Agustina. Verificado contra dos de sus cifras: $900.000 de montaje = US$596 y $24.500.000 de entelado = US$16.225. Editá esta celda y se recalcula todo el libro.'
ws['E4'].font = SUB

fila = 6
ws.cell(fila, 1, 'COSTO DEL EVENTO').font = Font(name=F, size=12, bold=True, color='1F3864')
fila += 1
lineas = [
 ('Costo total (cotizaciones reales + estimados)', f"='COSTOS UNIFICADOS'!J{fila_total}", USD0,
  'Suma de contratado, cotizado y lo que sólo tiene estimado. No incluye las alternativas descartadas.'),
 ('Lo que decía el presupuesto original', f"='COSTOS UNIFICADOS'!K{fila_total}", USD0,
  'Total de la hoja "Argentina (3 oct-4 oct)" del master: US$176.159.'),
 ('Diferencia contra el presupuesto', f"='COSTOS UNIFICADOS'!L{fila_total}", USD0,
  'Cuánto se corre el evento respecto de lo presupuestado, con los precios que hoy están sobre la mesa.'),
]
filas_kpi = {}
for etiqueta, formula, fmt, nota in lineas:
    ws.cell(fila, 1, etiqueta).font = NEGRO
    c = ws.cell(fila, 3, formula); c.font = VERDE; c.number_format = fmt; c.border = BOX
    ws.cell(fila, 5, nota).font = SUB
    filas_kpi[etiqueta] = fila
    fila += 1
f_total = filas_kpi['Costo total (cotizaciones reales + estimados)']
f_orig  = filas_kpi['Lo que decía el presupuesto original']
f_dif   = filas_kpi['Diferencia contra el presupuesto']
ws.cell(f_dif, 3).fill = FROJO

fila += 1
ws.cell(fila, 1, 'COSTO POR PERSONA').font = Font(name=F, size=12, bold=True, color='1F3864')
fila += 1
ws.cell(fila, 1, 'Base de cálculo').font = HDR; ws.cell(fila, 1).fill = FH
ws.cell(fila, 2, 'Personas').font = HDR; ws.cell(fila, 2).fill = FH
ws.cell(fila, 3, 'Costo por persona').font = HDR; ws.cell(fila, 3).fill = FH
ws.cell(fila, 5, 'Qué incluye').font = HDR; ws.cell(fila, 5).fill = FH
for j in (1, 2, 3, 5): ws.cell(fila, j).border = BOX
fila += 1
bases = [
 ('Aforo general', 5000, 'Asistentes de entrada general según la hoja del master.'),
 ('VIP', 1000, 'Asistentes VIP.'),
 ('Asistentes pagos (general + VIP)', 6000, 'La base más útil para decidir el precio de la entrada.'),
 ('Mastermind del martes', 246, 'Actividad aparte, no comparte la mayoría de los costos.'),
 ('Total de personas en el predio', 6108, 'Asistentes pagos + 100 de staff + 8 del equipo interno.'),
]
f_ini_bases = fila
for etiqueta, n, nota in bases:
    ws.cell(fila, 1, etiqueta).font = NEGRO
    c = ws.cell(fila, 2, n); c.font = AZUL; c.number_format = '#,##0'; c.fill = FIN
    c = ws.cell(fila, 3, f'=IFERROR($C${f_total}/B{fila},"")'); c.font = NEGRO; c.number_format = USD
    ws.cell(fila, 5, nota).font = SUB
    for j in (1, 2, 3): ws.cell(fila, j).border = BOX
    fila += 1
ws.cell(fila, 1, 'Las bases en amarillo son editables: cambiá el aforo y el costo por cabeza se recalcula.').font = SUB

fila += 2
ws.cell(fila, 1, 'ESTADO DE LOS RUBROS').font = Font(name=F, size=12, bold=True, color='1F3864')
fila += 1
for etiqueta, criterio, relleno, nota in [
    ('Contratado (hay contrato o anticipo)', 'Contratado', FVERD, 'Sede, las nueve facturas de merch e intercoms.'),
    ('Cotizado (hay presupuesto formal)', 'Cotizado', None, 'Sillas, WiFi, seguridad, servicio médico, limpieza, retiro de residuos y seguro.'),
    ('Sin cotizar (sólo estimado o nada)', 'Sin cotizar', FAMAR, 'Vuelos, alojamiento, ecobaños, marquetería, replanteo, unifilas, DJ, escoltas, Master Mind, personal logístico.'),
]:
    ws.cell(fila, 1, etiqueta).font = NEGRO
    c = ws.cell(fila, 3, f"=SUMIF('COSTOS UNIFICADOS'!$D:$D,\"{criterio}\",'COSTOS UNIFICADOS'!$J:$J)")
    c.font = VERDE; c.number_format = USD0; c.border = BOX
    if relleno: c.fill = relleno
    ws.cell(fila, 5, nota).font = SUB
    fila += 1
ws.cell(fila, 1, 'Alternativas descartadas (no suman)').font = SUB
c = ws.cell(fila, 3, "=SUMIF('COSTOS UNIFICADOS'!$D:$D,\"Alternativa\",'COSTOS UNIFICADOS'!$J:$J)")
c.font = SUB; c.number_format = USD0

# ============================================================================
# 3bis. MERCH — el detalle factura por factura
# ============================================================================
# (proveedor, item, cant, factura, fecha, ars_neto, iva, usdt_factura, pais,
#  pct_arg, pagado_pct, falta_regla, falta_factura, nota)
MERCH = [
 ('REMERASYESTAMPADOS', 'Pañuelos / pañoletas estampados', '8.500',
  '2601012 → saldo 2601019', '20/06 → 16/08', 13387500.0, 0.21, 5276.51,
  'Argentina + Uruguay', 0.65, 0.50, 5276.51, 5276.51,
  'La 2601019 ya viene emitida por el 50% restante. Falta definir con Uruguay como se reparte: aca se imputa el 65% a Argentina.'),
 ('REMERASYESTAMPADOS', 'Remeras premium estampadas', '250',
  '2601017', '16/08', 3527500.0, 0.21, 2780.64,
  'Argentina', 1.0, 0.0, 2780.64, 2780.64,
  'Se debe el 100%. Sube de 230 a 250 unidades contra la factura original 2601011, al mismo precio unitario de $14.110.'),
 ('REMERASYESTAMPADOS', 'Bordado computarizado en gorras', '940',
  '2601016', '10/08', 2643200.0, 0.0, 1721.95,
  'Argentina', 1.0, 0.0, 1721.95, 1721.95,
  'Se debe el 100%. Precio final sin IVA, con el descuento de $411.800 ya aplicado. Las gorras las provee el cliente.'),
 ('TEXTIL RYU S.R.L.', 'Gorra Flex unicolor/bicolor', '940',
  '2601009', '20/06', 2224980.0, 0.21, 0.0,
  'Argentina', 1.0, 1.0, 0.0, 0.0,
  'Pagada al 100% el 25/06. Es la compra de la gorra, distinta del bordado.'),
 ('LEOTEX', 'Lanyard premium doble raso 25 mm', '1.360',
  '2601013 → 2601018', '20/06 → 16/08', 2070600.0, 0.21, 1632.20,
  'Argentina', 1.0, 0.50, 816.10, 1632.20,
  'A CONFIRMAR: la carpeta se llama "Nuevo pago - 50%" pero la factura 2601018 viene por el total, igual que la original. El master registraba un saldo de US$850.'),
 ('DERQUI IMPRESIONES', 'Grafica oficial CMC — Argentina', '—',
  '2601014', '20/06 → 16/08', 3025260.0, 0.21, 2384.73,
  'Argentina', 1.0, 0.50, 1192.37, 2384.73,
  'A CONFIRMAR: la reemision del 16/08 viene por el total. El master registraba un saldo de US$1.839 al 13/09.'),
 ('DERQUI IMPRESIONES', 'Grafica oficial CMC — Uruguay', '—',
  '2601015', '20/06 → 16/08', 1564626.0, 0.21, 1233.35,
  'Uruguay', 0.0, 0.50, 616.68, 1233.35,
  'NO ES DE ARGENTINA. En la hoja PAGOS del master estaba cargada con pais = Argentina: es el monto que inflaba el total argentino.'),
 ('Blocko', 'Cintas portacredencial (alternativa descartada)', '1.300',
  'presupuesto 09/06', '09/06', 1749800.0, 0.21, 0.0,
  'Argentina', 0.0, 0.0, 0.0, 0.0,
  'Es el mismo item que los lanyards de LEOTEX, que ya esta contratado. Estaba sumando dos veces en la version anterior de esta planilla.'),
]

ws = hoja('MERCH', 'Todo el merch, factura por factura',
          'Carpeta de Drive "Presupuestos merch" · facturas 2601009 a 2601019 · emisor IIDAI LLC. Los proveedores son argentinos, no colombianos.')
COLS = [('Proveedor', 24), ('Ítem', 34), ('Cant.', 8), ('Factura', 20), ('Fecha', 15),
        ('Neto (ARS)', 15), ('IVA', 7), ('Total (ARS)', 15), ('USDT de la factura', 15),
        ('País', 18), ('% Argentina', 11), ('Costo Argentina (USD)', 15),
        ('Falta pagar · regla 50%', 15), ('Falta pagar · según factura', 15), ('Nota', 62)]
NC = len(COLS)
encabezados(ws, 4, COLS)
r = 5
filas_merch = []
for (prov, item, cant, fact, fecha, neto, iva, usdt, pais, pct, pagado, f_regla, f_fact, nota) in MERCH:
    ws.cell(r, 1, prov).font = BOLD
    ws.cell(r, 2, item).font = NEGRO
    ws.cell(r, 3, cant).font = NEGRO
    ws.cell(r, 4, fact).font = AZUL
    ws.cell(r, 5, fecha).font = AZUL
    c = ws.cell(r, 6, neto); c.font = AZUL; c.number_format = ARS
    c = ws.cell(r, 7, iva or None); c.font = AZUL; c.number_format = PCT
    c = ws.cell(r, 8, f'=F{r}*(1+N(G{r}))'); c.font = NEGRO; c.number_format = ARS
    c = ws.cell(r, 9, usdt or None); c.font = AZUL; c.number_format = USD
    ws.cell(r, 10, pais).font = NEGRO
    c = ws.cell(r, 11, pct); c.font = AZUL; c.number_format = PCT
    c = ws.cell(r, 12, f'=K{r}*H{r}/TABLERO!$C$4'); c.font = BOLD; c.number_format = USD
    c = ws.cell(r, 13, f'=K{r}*{f_regla}'); c.font = NEGRO; c.number_format = USD
    c = ws.cell(r, 14, f'=K{r}*{f_fact}'); c.font = NEGRO; c.number_format = USD
    ws.cell(r, 15, nota).font = SUB
    if pais == 'Uruguay' or pct == 0.0:
        for j in range(1, NC + 1): ws.cell(r, j).fill = FGRIS
    elif pagado == 0.0:
        for j in range(1, NC + 1): ws.cell(r, j).fill = FROJO
    elif pagado == 1.0:
        for j in range(1, NC + 1): ws.cell(r, j).fill = FVERD
    else:
        for j in range(1, NC + 1): ws.cell(r, j).fill = FAMAR
    for j in range(1, NC + 1):
        ws.cell(r, j).border = BOX
        ws.cell(r, j).alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[r].height = 46
    filas_merch.append(r)
    r += 1
ini_m, fin_m = filas_merch[0], filas_merch[-1]
ws.cell(r, 2, 'TOTAL IMPUTADO A ARGENTINA').font = Font(name=F, size=11, bold=True)
for col in (12, 13, 14):
    c = ws.cell(r, col, f'=SUM({L(col)}{ini_m}:{L(col)}{fin_m})')
    c.font = Font(name=F, size=11, bold=True); c.number_format = USD; c.fill = FVERD; c.border = BOX
fila_merch_total = r
ws.auto_filter.ref = f'A4:{L(NC)}4'

r += 2
ws.cell(r, 1, 'CÓMO LEER LOS COLORES').font = Font(name=F, size=11, bold=True, color='1F3864'); r += 1
for t, rel in [
    ('Verde — pagado al 100%: la compra de las gorras de Textil Ryu.', FVERD),
    ('Amarillo — 50% abonado, queda el saldo: pañuelos, lanyards y gráfica de Argentina.', FAMAR),
    ('Rojo — sin pagar nada, se debe el 100%: las remeras y el bordado de las gorras.', FROJO),
    ('Gris — no suma al costo de Argentina: la gráfica de Uruguay y la cotización descartada de Blocko.', FGRIS),
]:
    c = ws.cell(r, 1, t); c.font = NEGRO; c.fill = rel; c.border = BOX
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1

r += 1
ws.cell(r, 1, 'LO QUE IBA A COSTAR CONTRA LO QUE COSTÓ').font = Font(name=F, size=12, bold=True, color='1F3864'); r += 1
ws.cell(r, 1, 'Precios unitarios de merch.cmc2026.com al 06/07/2026, dólar $1.515. Compara el proveedor que traía Cumbre contra los proveedores argentinos que consiguió Agustina.').font = SUB; r += 1
CMP = [
 ('ARGENTINA', None, None, None, None, None),
 ('Remeras estampadas',            '100',   16.17, 1617.16,  9.31,  931.35),
 ('Gorras negras',                 '1.000',  9.90, 9900.99,  1.64, 1636.96),
 ('Cinta colgante / lanyard',      '1.000',  1.25, 1247.52,  1.00, 1004.95),
 ('Credenciales 10 × 13 cm',       '1.000',  0.57,  574.65,  0.48,  475.06),
 ('Hojas A4 (contrato)',           '5.500',  0.20, 1107.26,  0.05,  265.31),
 ('Tarjetas 21,5 × 10 cm (cheque)','5.500',  0.12,  671.62,  0.02,  109.78),
 ('Bolsas de friselina, 2 colores','1.000',  1.14, 1141.91,  0.64,  644.55),
 ('Pulseras tyvek — fuera de la comparación', '5.500', 0.10, 544.55, None, None),
 ('URUGUAY', None, None, None, None, None),
 ('Mapas de niveles de consciencia','3.000', 0.254, 763.00, 0.089, 266.14),
 ('Escarapelas / credenciales',     '600',   1.322, 793.00, 0.475, 285.03),
 ('Contratos (hojas A4)',           '3.000', 0.153, 458.00, 0.048, 144.71),
 ('Cheques (tarjetas)',             '3.000', 0.127, 382.00, 0.020,  59.88),
 ('Cordones con mosquetón — fuera de la comparación', '600', 1.50, None, None, None),
]
CC = [('Producto', 34), ('Cant.', 9), ('Unitario Cumbre', 13), ('Total Cumbre', 13),
      ('Unitario Agustina', 13), ('Total Agustina', 13), ('Ahorro', 13), ('Ahorro %', 10)]
for j, (t, an) in enumerate(CC, 1):
    c = ws.cell(r, j, t); c.font = HDR; c.fill = FH; c.border = BOX
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
r += 1
ini_c = r
ini_pais, filas_pais, subtotales_pais = r, [], []
for prod, cant, uc, tc, ua, ta in CMP:
    if cant is None:                      # banda de pais
        if filas_pais:
            subtotales_pais.append((pais_actual, ini_pais, r - 1))
            ws.cell(r, 1, f'Subtotal {pais_actual}').font = BOLD
            for col, letra in ((4, 'D'), (6, 'F'), (7, 'G')):
                c = ws.cell(r, col, f'=SUMIF($F${ini_pais}:$F${r-1},"<>",{letra}{ini_pais}:{letra}{r-1})')
                c.font = BOLD; c.number_format = USD; c.fill = FGRIS; c.border = BOX
            c = ws.cell(r, 8, f'=IFERROR(G{r}/D{r},"")'); c.font = BOLD; c.number_format = PCT
            c.fill = FGRIS; c.border = BOX
            for j in range(1, 9): ws.cell(r, j).border = BOX
            r += 1
            filas_pais = []
        r = banda_bloque(ws, r, prod, 8)
        pais_actual, ini_pais = prod.capitalize(), r
        continue
    filas_pais.append(r)
    ws.cell(r, 1, prod).font = BOLD if ua is not None else SUB
    ws.cell(r, 2, cant).font = NEGRO
    c = ws.cell(r, 3, uc); c.font = AZUL; c.number_format = USD
    c = ws.cell(r, 4, tc); c.font = AZUL; c.number_format = USD
    c = ws.cell(r, 5, ua); c.font = AZUL; c.number_format = USD
    c = ws.cell(r, 6, ta); c.font = AZUL; c.number_format = USD
    if ua is not None:
        c = ws.cell(r, 7, f'=D{r}-F{r}'); c.font = BOLD; c.number_format = USD; c.fill = FVERD
        c = ws.cell(r, 8, f'=G{r}/D{r}'); c.font = NEGRO; c.number_format = PCT
    else:
        for j in range(1, 9): ws.cell(r, j).fill = FGRIS
    for j in range(1, 9):
        ws.cell(r, j).border = BOX
    r += 1
if filas_pais:
    subtotales_pais.append((pais_actual, ini_pais, r - 1))
    ws.cell(r, 1, f'Subtotal {pais_actual}').font = BOLD
    for col, letra in ((4, 'D'), (6, 'F'), (7, 'G')):
        c = ws.cell(r, col, f'=SUMIF($F${ini_pais}:$F${r-1},"<>",{letra}{ini_pais}:{letra}{r-1})')
        c.font = BOLD; c.number_format = USD; c.fill = FGRIS; c.border = BOX
    c = ws.cell(r, 8, f'=IFERROR(G{r}/D{r},"")'); c.font = BOLD; c.number_format = PCT
    c.fill = FGRIS; c.border = BOX
    for j in range(1, 9): ws.cell(r, j).border = BOX
    r += 1
fin_c = r - 1
ws.cell(r, 1, 'TOTAL DE LA CANASTA — ARGENTINA + URUGUAY').font = Font(name=F, size=11, bold=True)
suma_sub = {col: '+'.join(f'{letra}{fin}' for _, _, fin in
            [(a, b, c2 + 1) for a, b, c2 in subtotales_pais]) for col, letra in
            ((4, 'D'), (6, 'F'), (7, 'G'))}
for col, letra in ((4, 'D'), (6, 'F'), (7, 'G')):
    c = ws.cell(r, col, '=' + '+'.join(f'{letra}{fin + 1}' for _, _, fin in subtotales_pais))
    c.font = Font(name=F, size=11, bold=True); c.number_format = USD; c.fill = FVERD; c.border = BOX
c = ws.cell(r, 8, f'=G{r}/D{r}'); c.font = Font(name=F, size=11, bold=True); c.number_format = PCT; c.fill = FVERD; c.border = BOX
r += 2
for t in [
 'La gorra sola explica la mitad del ahorro: el proveedor de Cumbre la cobraba US$9,90 y Agustina la consiguió a US$1,64. Sobre 1.000 unidades son US$8.264.',
 'Las pulseras tyvek quedan fuera de los dos lados de la cuenta porque no hay precio comparable del lado argentino.',
 'OJO: esta canasta NO es el total del rubro merch. Los pañuelos (US$6.973) y el bordado de gorras (US$1.750) no entran en la comparación, y las cantidades',
 'del sitio no son exactamente las de las facturas finales (remeras 100 contra 250, gorras 1.000 contra 940). Sirve para medir el ahorro por precio unitario,',
 'no para reemplazar el costo del rubro, que sale de las nueve facturas de arriba.',
 'Uruguay va con su propio subtotal: US$2.396,00 contra US$755,77, US$1.640,23 de ahorro. Es la misma negociación con los mismos proveedores, pero NO entra en el costo del evento de Argentina.',
 'Los cordones con mosquetón de Uruguay quedan fuera de la cuenta por la misma razón que las pulseras tyvek: no hay precio comparable del lado argentino.',
 'Sumando los dos países, la canasta pasa de US$18.657,12 a US$5.823,74: US$12.833,38 de ahorro, un 68,8%.',
]:
    ws.cell(r, 1, t).font = SUB
    r += 1

r += 1
ws.cell(r, 1, 'TRES COSAS QUE HAY QUE CONFIRMAR').font = Font(name=F, size=11, bold=True, color='C00000'); r += 1
for t in [
 '1) Lanyards y gráfica de Argentina: las facturas reemitidas en agosto vienen por el TOTAL, pero la carpeta y el master dicen que ya se abonó el 50%. Si las facturas están bien, hay que pagar US$2.008 más de lo que sale por la regla del 50%. Una sola consulta a LEOTEX y a Derqui lo resuelve.',
 '2) Pañuelos: son 8.500 unidades compartidas entre Argentina y Uruguay y la factura no las separa. Acá se imputó el 65% a Argentina (5.500 de 8.500, la misma proporción que la gráfica). Si van todos a Argentina, el costo sube US$3.755.',
 '3) La gráfica de Uruguay (US$1.233) estaba cargada en el master como Argentina. Ya quedó fuera del costo argentino, pero conviene corregirla también en el master para que no se arrastre.',
]:
    c = ws.cell(r, 1, t); c.font = NEGRO
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=15)
    c.alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[r].height = 30
    r += 1

# ============================================================================
# 3ter. PAGOS — cuanto se pago, cuanto falta y cuando
# ============================================================================
PAGADO = [
 ('Sede',       'La Rural — anticipo del contrato', '27/04/2026', 43446.0, 'Hoja PAGOS del master'),
 ('Producción', 'Intercoms — 100%',                 'ya comprados',  560.0, 'Confirmado por Agustina el 03/09'),
 ('Técnica',    'Grupo MET — factura INV-11',       '31/07/2026',  5000.0, 'Recibo de Mercury del 31/07'),
 ('Merch',      'Pañuelos — 50% (parte Argentina)', '24/06/2026',  3569.15, 'Master: 50% de US$10.982, imputado al 65%'),
 ('Merch',      'Gorras — 100%',                    '25/06/2026',  1825.0, 'Factura 2601009 de Textil Ryu'),
 ('Merch',      'Lanyards — 50%',                   '25/06/2026',   849.0, 'Hoja PAGOS del master'),
 ('Merch',      'Gráfica Argentina — anticipo',     '25/06/2026',   642.0, 'Master: total US$2.481, saldo US$1.839'),
]
CUOTAS = [
 ('Sede',       'La Rural — saldo, cuota 1 de 2', '24/08/2026', 15777.0, 'VENCIDA',   'Cronograma del master'),
 ('Merch',      'Remeras (250)',                  'ya facturado', 2780.64, 'SIN PAGAR', 'Factura 2601017 del 16/08 · NO figura en el master'),
 ('Merch',      'Bordado de gorras (940)',        'ya facturado', 1721.95, 'SIN PAGAR', 'Factura 2601016 del 10/08 · NO figura en el master'),
 ('Merch',      'Pañuelos — saldo (parte Argentina)', '13/09/2026', 3429.73, 'Programado', 'Factura 2601019 · 65% de 5.276,51 USDT'),
 ('Merch',      'Gráfica Argentina — saldo',      '13/09/2026',  1192.37, 'Programado', 'Master. La factura 2601014 viene por el total: a confirmar'),
 ('Merch',      'Lanyards — saldo',               '13/09/2026',   816.10, 'Programado', 'Master. La factura 2601018 viene por el total: a confirmar'),
 ('Sede',       'La Rural — saldo, cuota 2 de 2', '23/09/2026', 15777.0, 'Programado', 'Cronograma del master'),
]

ws = hoja('PAGOS', 'Cuánto se pagó, cuánto falta y cuándo',
          'Sólo lo que está documentado. Lo que no tiene fecha acordada va abajo, en el resto.')
COLS = [('Bloque', 14), ('Concepto', 42), ('Fecha', 15), ('USD', 14), ('Estado', 14), ('De dónde sale', 58)]
NC = len(COLS)
encabezados(ws, 4, COLS)
r = 5
r = banda(ws, r, 'YA PAGADO', NC)
ini_p = r
for bl, con, fe, mo, fu in PAGADO:
    ws.cell(r, 1, bl).font = SUB
    ws.cell(r, 2, con).font = NEGRO
    ws.cell(r, 3, fe).font = AZUL
    c = ws.cell(r, 4, mo); c.font = BOLD; c.number_format = USD
    c = ws.cell(r, 5, 'Pagado'); c.font = NEGRO
    ws.cell(r, 6, fu).font = SUB
    for j in range(1, NC + 1): ws.cell(r, j).fill = FVERD; ws.cell(r, j).border = BOX
    r += 1
fila_pag = r
ws.cell(r, 2, 'Pagado a hoy').font = BOLD
c = ws.cell(r, 4, f'=SUM(D{ini_p}:D{r-1})'); c.font = BOLD; c.number_format = USD
for j in range(1, NC + 1): ws.cell(r, j).fill = FGRIS; ws.cell(r, j).border = BOX
r += 2

r = banda(ws, r, 'FALTA PAGAR · CON FECHA', NC)
ini_c = r
for bl, con, fe, mo, st, fu in CUOTAS:
    ws.cell(r, 1, bl).font = SUB
    ws.cell(r, 2, con).font = NEGRO
    ws.cell(r, 3, fe).font = AZUL
    c = ws.cell(r, 4, mo); c.font = BOLD; c.number_format = USD
    ws.cell(r, 5, st).font = BOLD
    ws.cell(r, 6, fu).font = SUB
    relleno = FROJO if st in ('VENCIDA', 'SIN PAGAR') else None
    for j in range(1, NC + 1):
        if relleno: ws.cell(r, j).fill = relleno
        ws.cell(r, j).border = BOX
    r += 1
fila_cuo = r
ws.cell(r, 2, 'Comprometido con fecha').font = BOLD
c = ws.cell(r, 4, f'=SUM(D{ini_c}:D{r-1})'); c.font = BOLD; c.number_format = USD
for j in range(1, NC + 1): ws.cell(r, j).fill = FGRIS; ws.cell(r, j).border = BOX
r += 1
ws.cell(r, 2, 'De eso, vencido o facturado sin pagar').font = SUB
c = ws.cell(r, 4, f'=SUMIF(E{ini_c}:E{fila_cuo-1},"VENCIDA",D{ini_c}:D{fila_cuo-1})'
                  f'+SUMIF(E{ini_c}:E{fila_cuo-1},"SIN PAGAR",D{ini_c}:D{fila_cuo-1})')
c.font = BOLD; c.number_format = USD; c.fill = FROJO; c.border = BOX
r += 2

r = banda(ws, r, 'RESUMEN', NC)
for etiqueta, formula, relleno in [
    ('Costo total del evento', f"='COSTOS UNIFICADOS'!J{fila_total}", FGRIS),
    ('Ya pagado',              f'=D{fila_pag}',                        FVERD),
    ('Falta desembolsar',      f"='COSTOS UNIFICADOS'!J{fila_total}-D{fila_pag}", FAMAR),
    ('   de eso, con fecha acordada', f'=D{fila_cuo}',                 None),
    ('   de eso, todavía sin fecha',  f"='COSTOS UNIFICADOS'!J{fila_total}-D{fila_pag}-D{fila_cuo}", None)]:
    ws.cell(r, 2, etiqueta).font = BOLD if not etiqueta.startswith('  ') else NEGRO
    c = ws.cell(r, 4, formula); c.font = BOLD; c.number_format = USD0; c.border = BOX
    if relleno: c.fill = relleno
    r += 1
r += 1
for t in [
 'Las remeras y el bordado de gorras están facturados desde agosto y no figuran en el cronograma del master: son US$4.503 que hoy no están en ningún calendario.',
 'El resto sin fecha son los rubros cerrados y cotizados que todavía no tienen forma de pago acordada: técnica, entelado, CCTV, catering, sillas y los servicios del predio.',
 'Las facturas reemitidas de LEOTEX y Derqui vienen por el total aunque el master registra el 50% abonado. Acá se tomó la regla del 50%: si las facturas están bien, son US$2.008 más.',
]:
    ws.cell(r, 2, t).font = SUB
    r += 1

# ============================================================================
# 4. DECISIONES ABIERTAS
# ============================================================================
ws = hoja('DECISIONES', 'Decisiones abiertas que mueven el número',
          'Cada fila cambia el costo total. Ordenadas por cuánto pesan.')
COLS = [('Decisión', 34), ('Opción A', 34), ('Opción B', 34), ('Diferencia', 15), ('Qué hay que hacer', 56)]
encabezados(ws, 4, COLS)
DEC = [
 ('Infraestructura de La Rural',
  'Presupuesto completo: $66.494.708 (US$43.461), con las 5.000 sillas y todo el mobiliario',
  'Sólo panelería + dirección técnica + guardias: $16.858.366 (US$11.019). El mobiliario se compra aparte.',
  '=(66494708.2-16858366.2)/TABLERO!$C$4',
  'Agustina marca ítem por ítem qué se queda. La hoja INFRAESTRUCTURA tiene las 28 líneas del presupuesto para filtrar.'),
 ('Montaje de las sillas',
  'La Rural con las sillas incluidas y armadas',
  'FDL Eventos + montaje/acomodación/desmontaje negociado en $900.000',
  '', 'Con FDL a $31.218.000 más $900.000 de montaje son US$20.992 contra US$27.778 de La Rural. Cerrar el número del montaje y firmar.'),
 ('Catering — cerrado', 'Grupo Ambient: US$14.569 por las dos propuestas',
  'La planilla tenía US$33.820 estimados',
  '=14569-33820', 'Cerrado: 600 desayunos VIP (no 800) + comidas para 150 personas. '
  'US$19.251 por debajo de lo presupuestado. OJO: la factura 2601021 está calculada sobre 800 desayunos; '
  'pedirle a Ángeles que la reemita por 600 antes de pagar el anticipo.'),
 ('Rubros nuevos sin cotizar',
  'Ecobaños, marquetería de certificaciones y premios, replanteo de sillas con ingeniero',
  'Hoy entran en US$0',
  '', 'Pedir las tres cotizaciones esta semana. Son los últimos huecos que quedan además de vuelos y alojamiento.'),
 ('Merch — el 50% que falta pagar',
  'Regla del 50%: quedan US$9.941 por transferir en USDT',
  'Según las facturas reemitidas en agosto: US$11.949',
  '=11949.25-9940.79',
  'Las facturas de LEOTEX (lanyards) y Derqui (gráfica ARG) vienen por el total aunque ya se abonó el 50%. Una consulta a cada proveedor cierra la diferencia. Lo urgente igual son las remeras (US$2.781) y el bordado de gorras (US$1.722): esos no tienen nada pagado.'),
 ('Merch — cómo se reparten los pañuelos con Uruguay',
  'Los 8.500 pañuelos se imputan 65% a Argentina (US$6.973)',
  'Si van todos a Argentina: US$10.728',
  '=0.35*13387500*1.21/TABLERO!$C$4',
  'La factura no separa las unidades por país. Definir el reparto con Uruguay antes de cerrar el número para el CEO.'),
 ('Vuelos y alojamiento del equipo',
  'Sin cotizar', 'Sin cotizar', '',
  'No hay ni un mail ni una línea en el sheet. Con 8 personas de equipo interno, es el hueco más grande que queda.'),
 ('Facturación al exterior',
  'FDL Eventos y Vittal no confirmaron si pueden facturar a la empresa de EE.UU.',
  'Buscar proveedores que sí facturen, o resolver un circuito de pago local',
  '', 'Resolverlo antes de firmar: puede trabar los dos contratos.'),
]
r = 5
for dec, a, b, formula, accion in DEC:
    ws.cell(r, 1, dec).font = BOLD
    ws.cell(r, 2, a).font = NEGRO
    ws.cell(r, 3, b).font = NEGRO
    c = ws.cell(r, 4, formula or None); c.font = NEGRO; c.number_format = USD
    if formula: c.fill = FAMAR
    ws.cell(r, 5, accion).font = NEGRO
    for j in range(1, 6):
        ws.cell(r, j).border = BOX
        ws.cell(r, j).alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[r].height = 56
    r += 1

# ============================================================================
# 6. AHORROS — comparativa entre lo cotizado y lo que se cerró
# ============================================================================
# (rubro, referencia, ref_ars, ref_usd, cerrado, cer_ars, cer_usd, respaldo)
# ref_ars/cer_ars en None cuando el número está directamente en dólares.
AHORROS = [
 ('Infraestructura — armado de stands',
  'La Rural: panelería + dirección técnica + guardias', 16858366.2, None,
  'Armado del stand de staff, el de producción y el camarín', None, 9270.0,
  'Presupuesto de La Rural del 17/08 (mail criccio@larural.com.ar): panelería $9.938.187 más generales $6.920.179. Cerrado por Agustina en US$9.270.'),
 ('Mobiliario — sillas y montaje',
  'La Rural: 5.000 sillas imperio bordó, armado incluido', None, 28431.0,
  'FDL Eventos: 5.500 sillas + flete, más el montaje, la acomodación y el desmontaje', None, 17450.0,
  'Cifras de Agustina, sin IVA. Reemplaza la fila que estaba aparte como "sillas": es la misma negociación, estaba contada dos veces. Verificado: 17.450 al dólar de referencia son $26.698.500, que es el presupuesto FDL 026-2697 ($25.800.000) más los $900.000 del montaje. Agustina tiene los presupuestos anteriores para adjuntar.'),
 ('Entelado',
  'Primer presupuesto recibido', None, 25490.0,
  'Negociado por Agustina', None, 16500.0,
  'Cifras de Agustina, sin IVA. Verificado: 25.490 al dólar de referencia son $39.000.000 exactos, que es la primera cotización. Agustina tiene el presupuesto anterior para adjuntar.'),
 ('Catering',
  'Primera propuesta de Grupo Ambient', 30000000.0, None,
  '600 lunchbox VIP + desayuno, almuerzo y cena para 150 personas', None, 14569.0,
  'Cierre informado por Agustina. En el correo está la propuesta de AmbientHouse del 16/06 a $28.500 + IVA por persona por día con base mínima de 480.'),
 ('Merch e impresos',
  'Proveedor de Cumbre, precios unitarios al 06/07', None, 16261.12,
  'Proveedores argentinos conseguidos por Agustina', None, 5067.97,
  'Comparativa de merch.cmc2026.com al 06/07/2026 (dólar $1.515). Es una canasta cerrada de siete ítems: remeras, gorras, lanyards, credenciales, hojas A4, tarjetas y bolsas de friselina. Las pulseras tyvek quedan fuera de los dos lados porque no hay precio comparable. NO es el total del rubro merch: los pañuelos y el bordado de gorras no entran en esta comparación.'),
 ('Vallado del público',
  'Presupuestado en la hoja de Argentina', None, 1250.0,
  '500 vallas bonificadas por el proveedor de técnica', None, 0.0,
  'Negociación de Agustina. Además bonifican los efectos especiales, que no estaban valorizados.'),
]

ws = hoja('AHORROS', 'Lo cotizado contra lo que se cerró',
          'Cuánto se bajó rubro por rubro. La columna "Respaldo" dice si la referencia está documentada en el correo o si es un dato aportado.')
COLS = [('Rubro', 26), ('Referencia cotizada', 40), ('Ref. en pesos', 15), ('Ref. USD', 12),
        ('Lo que se cerró', 44), ('Cerrado en pesos', 15), ('Cerrado USD', 12),
        ('Ahorro USD', 13), ('Ahorro %', 10), ('Respaldo', 60)]
NC = len(COLS)
encabezados(ws, 4, COLS)
r = 5
filas_ah = []
for rubro, ref, ref_ars, ref_usd, cer, cer_ars, cer_usd, respaldo in AHORROS:
    ws.cell(r, 1, rubro).font = BOLD
    ws.cell(r, 2, ref).font = NEGRO
    c = ws.cell(r, 3, ref_ars); c.font = AZUL; c.number_format = ARS
    c = ws.cell(r, 4, ref_usd if ref_usd is not None else f'=C{r}/TABLERO!$C$4')
    c.font = AZUL if ref_usd is not None else NEGRO; c.number_format = USD
    ws.cell(r, 5, cer).font = NEGRO
    c = ws.cell(r, 6, cer_ars); c.font = AZUL; c.number_format = ARS
    c = ws.cell(r, 7, cer_usd if cer_usd is not None else f'=F{r}/TABLERO!$C$4')
    c.font = AZUL if cer_usd is not None else NEGRO; c.number_format = USD
    c = ws.cell(r, 8, f'=D{r}-G{r}'); c.font = BOLD; c.number_format = USD; c.fill = FVERD
    c = ws.cell(r, 9, f'=IFERROR(H{r}/D{r},"")'); c.font = NEGRO; c.number_format = PCT
    ws.cell(r, 10, respaldo).font = SUB
    for j in range(1, NC + 1):
        ws.cell(r, j).border = BOX
        ws.cell(r, j).alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[r].height = 46
    filas_ah.append(r)
    r += 1
ws.cell(r, 1, 'AHORRO TOTAL').font = Font(name=F, size=12, bold=True)
for col in (4, 7, 8):
    c = ws.cell(r, col, f'=SUM({L(col)}{filas_ah[0]}:{L(col)}{filas_ah[-1]})')
    c.font = Font(name=F, size=12, bold=True); c.number_format = USD; c.fill = FVERD; c.border = BOX
fila_ah_total = r

r += 2
ws.cell(r, 1, 'TÉCNICA — la comparación no es directa').font = Font(name=F, size=11, bold=True, color='C00000')
r += 1
for t in [
 'Se pidió cotización de técnica a NUEVE empresas el 13 y 14 de junio. Respondieron seis con número. El detalle completo está en la hoja COTIZACIONES TÉCNICA.',
 '',
 '   · Sound-Light (16/06): el PDF no se pudo abrir, el mail pesa 33 MB. El mail del 19/06 les dice que las otras estaban "prácticamente la mitad", lo que lo ubica por encima de los $100.000.000.',
 '   · Bonetto (17/06): $94.936.000 — sonido e iluminación $69.736.000, LED $20.720.000, CCTV $4.480.000. Grúas y efectos aparte (la chispa fría la cobra POR MINUTO).',
 '   · Prina (18/06): $48.885.000 + IVA = $59.150.850. Pero cotiza 40 m² de LED cuando el rider pide 65, y deja afuera grúas y energía. Vencida.',
 '   · 2MG (24/06): $45.000.000 con descuento + IVA = $54.450.000, más $19.202.700 de adicionales. Sonido más liviano, cotizado para "2000 / 5000 pax".',
 '   · Dixi Group (26/06): $82.000.000 + IVA = $99.220.000 llave en mano, con rigging completo, encomiendas y 4 máquinas de chispa fría.',
 '   · VMG (29/06): $15.276.250 con IVA, sólo las pantallas LED de 65 m².',
 '',
 'POR QUÉ TÉCNICA NO ENTRA EN LA TABLA DE ARRIBA: Grupo MET cerró en US$60.000 ($90.600.000) y con el circuito cerrado el paquete queda cerca de $100.000.000.',
 'Eso es prácticamente lo mismo que Dixi llave en mano y menos que Bonetto, pero está POR ENCIMA de 2MG y de Prina. Poner esa diferencia como ahorro no se sostendría',
 'delante del CEO, porque los alcances no son iguales: Prina cotizó 38% menos de LED, 2MG un sonido más chico, y Grupo MET incluye lo que se agregó tras la visita al predio.',
 '',
 'LO QUE SÍ SE SOSTIENE EN TÉCNICA: las 500 vallas y los efectos especiales van bonificados. Bonetto los cobra: $300.000 la máquina de humo y $600.000 POR MINUTO las dos de chispa fría.',
 'Con los 6 disparos por día que pide el rider, en cualquiera de las otras propuestas eso es plata.',
 '',
 'FALTA UNA SOLA COSA para cerrar el ahorro de técnica con un número: el PDF de Grupo MET con el alcance final. Es el único de los siete proveedores sin presupuesto escrito en el correo.',
]:
    ws.cell(r, 1, t).font = SUB
    r += 1

# ============================================================================
# 6bis. COTIZACIONES DE TÉCNICA — las siete propuestas que llegaron
# ============================================================================
# (proveedor, fecha, contacto, sin_iva, con_iva, alcance, aparte, respaldo)
# sin_iva/con_iva en None cuando no hay número firme.
TECNICA = [
 ('Sound-Light', '16/06/2026', 'ventas@sound-light.com.ar (Noelia Escudero)',
  None, None,
  'Técnica integral sobre el pliego base.',
  'Sin detalle: no se pudo abrir el adjunto.',
  'El mail pesa 33 MB y el PDF no se puede extraer desde acá. Lo que sí está documentado es el mail del 19/06 en el que Agustina les dice que las otras cotizaciones estaban "prácticamente la mitad de lo cotizado por ustedes". Sobre las de esa fecha (Bonetto y Prina) eso lo ubica por encima de los $100.000.000.'),
 ('Bonetto Sonido e Iluminación', '17/06/2026', 'cristian@bonetto.net',
  94936000.0, None,
  'Sonido e iluminación $69.736.000 · pantallas LED $20.720.000 · CCTV $4.480.000. 16 L-Acoustics Kiva II, 240 LED ST18-15, 322 truss box, consola Yamaha QL5.',
  'Grúas (2 autoelevadores tijera x 3 días) $4.480.000 · máquina de humo $300.000 · 2 máquinas de chispa fría $600.000 POR MINUTO · minuto adicional $150.000.',
  'Presupuesto PR26-287 V.1. El PDF viene con el texto vectorizado: hubo que renderizarlo a imagen para leerlo. No aclara si los valores llevan IVA. Dice expresamente "NO CONTEMPLA RIDER DE BANDA".'),
 ('Prina', '18/06/2026', 'valentin.andina@prina.net',
  48885000.0, 59150850.0,
  'LED 40 m² $6.700.000 · CCTV 2 cámaras $4.510.000 · sonido (16 Adamson Y10) $10.465.000 · iluminación $12.490.000 · RRHH y logística $14.720.000.',
  'Incluye 2 máquinas Sparkular y máquina de humo.',
  'OJO CON EL ALCANCE: cotiza 40 m² de LED (6×4 + 2 de 2×4) cuando el rider pide 65 m² (7×5 + 2 de 3×5). No incluye grúas, brazos articulados ni tijeras, ni la generación de energía. Validez de 10 días: vencida desde el 28/06.'),
 ('2MG', '24/06/2026', 'rocio.g@2mg.net',
  45000000.0, 54450000.0,
  'Sonido, video (LED 65 m², medida correcta), iluminación, CCTV, efectos especiales, RRHH y logística. Total de lista $53.655.470, con descuento especial $45.000.000.',
  'Extras de iluminación $10.177.500 · escenario 13×3 m con pasarela $5.692.500. Los dos juntos suman $15.870.000 sin IVA.',
  'Incluye certificación de planos por escribano, eléctrico matriculado y autoelevadora, hasta 8 puntos de colgado. El sonido es más liviano que el de los otros (12 gabinetes DVA T8) y el presupuesto dice "2000 / 5000 pax". Validez 15 días.'),
 ('Dixi Group', '26/06/2026', 'g.judcovski@dixieventos.com.ar',
  82000000.0, 99220000.0,
  'Llave en mano: sonido, iluminación (90 par LED, 6 beam, 8 wash), video (LED 65 m²), CCTV y rigging completo (6 puentes de 12 m, 14 aparejos de 1 T, motores).',
  'Nada: es llave en mano.',
  'Es el alcance más completo de todos. Incluye 4 máquinas de chispa fría con 8 disparos por día, plano general con puntos de colgado firmado por arquitecto, encomienda eléctrica de profesional matriculado y productores generales de técnica.'),
 ('VMG Visual Solutions', '29/06/2026', 'presupuestos@vmg-web.com',
  12625000.0, 15276250.0,
  'Sólo pantallas LED: 65 m² (7×5 central + 2 de 3×5), pitch 2.6 indoor blackface, con control, servidor y operador.',
  'Todo lo demás.',
  'NO ES COMPARABLE con las otras: es únicamente video. Sirve para poner precio al renglón de LED. No incluye estructuras de colgado, rigging, tarimas, entelados ni autoelevadores. Recotiza si el dólar salta más de 15%.'),
 ('Grupo MET — ES EL QUE SE CONTRATÓ', 'sin PDF', 'lalo@somosgrupomet.com',
  None, 90600000.0,
  'Sonido, iluminación, video y LED por US$60.000. Aparte: entelado $24.500.000 y circuito cerrado US$6.209.',
  'Bonifica 500 vallas y los efectos especiales, que en las otras cotizaciones se pagan.',
  'ES EL ÚNICO DE LOS SIETE SIN PRESUPUESTO ESCRITO EN EL CORREO. Se revisó todo el Gmail: en el hilo con Lalo Aizenberg hay ocho idas y vueltas entre el 16 y el 30 de junio y una visita al predio el 30/06, pero nunca llegó un PDF. Lo único que hay es la factura INV-11 de Mercury por US$5.000, pagada el 31/07.'),
]

ws = hoja('COTIZACIONES TÉCNICA', 'Las siete propuestas de técnica que llegaron',
          'Se pidió cotización a nueve empresas entre el 13 y el 14 de junio. Respondieron seis con número; Black-Out y 4A Latam nunca contestaron y una dirección rebotó.')
COLS = [('Proveedor', 30), ('Fecha', 12), ('Contacto', 34), ('Sin IVA', 16), ('Con IVA 21%', 16),
        ('USD al dólar de trabajo', 15), ('Qué incluye', 62), ('Qué queda aparte', 44), ('Observaciones', 70)]
NC = len(COLS)
encabezados(ws, 4, COLS)
r = 5
for prov, fecha, contacto, sin_iva, con_iva, alcance, aparte, respaldo in TECNICA:
    ws.cell(r, 1, prov).font = BOLD
    ws.cell(r, 2, fecha).font = AZUL
    ws.cell(r, 3, contacto).font = SUB
    c = ws.cell(r, 4, sin_iva); c.font = AZUL; c.number_format = ARS
    c = ws.cell(r, 5, con_iva if con_iva is not None else (f'=D{r}*1.21' if sin_iva else None))
    c.font = AZUL if con_iva is not None else NEGRO; c.number_format = ARS
    c = ws.cell(r, 6, f'=IF(E{r}="","",E{r}/TABLERO!$C$4)'); c.font = NEGRO; c.number_format = USD0
    ws.cell(r, 7, alcance).font = NEGRO
    ws.cell(r, 8, aparte).font = NEGRO
    ws.cell(r, 9, respaldo).font = SUB
    if prov.startswith('Grupo MET'):
        for j in range(1, NC + 1): ws.cell(r, j).fill = FVERD
    elif prov.startswith('VMG'):
        for j in range(1, NC + 1): ws.cell(r, j).fill = FGRIS
    elif sin_iva is None:
        for j in range(1, NC + 1): ws.cell(r, j).fill = FAMAR
    for j in range(1, NC + 1):
        ws.cell(r, j).border = BOX
        ws.cell(r, j).alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[r].height = 86
    r += 1

r += 1
ws.cell(r, 1, 'CÓMO LEER ESTA TABLA — Y POR QUÉ NO SE PUEDE RESTAR SIN MÁS').font = Font(name=F, size=11, bold=True, color='C00000')
r += 1
for t in [
 'Los siete números NO cotizan lo mismo, así que la diferencia contra Grupo MET no es un ahorro hasta que se iguale el alcance:',
 '   · Grupo MET cerró en US$60.000 (unos $90.600.000) por sonido, iluminación, video y LED, con todo lo que se fue agregando después de la visita al predio del 30 de junio.',
 '   · Sumándole el circuito cerrado (US$6.209) el paquete de técnica queda cerca de los $100.000.000, que es prácticamente lo mismo que cotizó Dixi llave en mano ($99.220.000) y por debajo de Bonetto ($114.872.560 más grúas y efectos).',
 '   · En cambio queda POR ENCIMA de 2MG con descuento ($54.450.000 más $19.202.700 de adicionales = $73.652.700) y de Prina ($59.150.850).',
 '   · Pero Prina cotizó 40 m² de LED cuando el rider pide 65, y dejó afuera las grúas y la energía; y 2MG cotizó un sonido más liviano, para "2000 / 5000 pax".',
 '',
 'Lo que sí está documentado y se sostiene solo:',
 '   · Sound-Light quedó afuera por precio, y hay un mail del 19/06 pidiéndoles que revisen porque las otras estaban "prácticamente la mitad".',
 '   · Los efectos especiales y las 500 vallas los bonifica Grupo MET. Bonetto los cobra: $300.000 la máquina de humo y $600.000 POR MINUTO las dos de chispa fría.',
 '   · Se pidieron nueve cotizaciones y se compararon seis. Ese trabajo está entero en el correo.',
 '',
 'PARA CERRAR EL AHORRO DE TÉCNICA FALTA UNA SOLA COSA: el PDF de Grupo MET con el alcance final. Con eso se puede comparar contra el mismo alcance en Dixi, Bonetto, 2MG y Prina, y recién ahí poner un número.',
]:
    ws.cell(r, 1, t).font = SUB if not t.startswith('PARA CERRAR') else Font(name=F, size=10, bold=True, color='C00000')
    r += 1

# ============================================================================
# 5. INFRAESTRUCTURA — las 28 líneas de La Rural, para marcar cuáles se quedan
# ============================================================================
INFRA = [
 ('PANELERÍA (montaje de estructuras)', [
   ('Oficina de producción — 25 m² (panelería + puerta + iluminación)', 1314575.00),
   ('Sala de staff — 100 m² (panelería + puerta + iluminación)',        5258300.00),
   ('Depósito 8×8 para las 2.500 barras — 64 m²',                       3365312.00)]),
 ('MOBILIARIO · oficina de producción', [
   ('Mesa redonda 1,60 ×2', 236000.0), ('Silla imperio bordó ×10', 85000.0),
   ('Sillón Valencia 2C ×2', 254000.0), ('Sillón BRNO ×2', 160000.0),
   ('Mesa ratona ×1', 118000.0), ('Perchero de pie ×1', 40000.0),
   ('Tacho de basura grande ×1', 56300.0)]),
 ('MOBILIARIO · sala de staff', [
   ('Mesa redonda 1,60 ×5', 590000.0), ('Silla imperio bordó ×40', 340000.0),
   ('Perchero de pie ×1', 40000.0), ('Tacho de basura grande ×1', 56300.0)]),
 ('MOBILIARIO · sala principal', [
   ('Silla imperio bordó ×5.000', 42500000.0), ('Mesa 1,80 × 0,90 ×40', 2500000.0),
   ('Mantel ×40', 900000.0), ('Mantel redondo ×8', 256000.0)]),
 ('MOBILIARIO · camarín de speakers', [
   ('Mesa redonda 1,60 ×1', 118000.0), ('Espejo ×1', 79500.0),
   ('Juego de living (sillones + mesa ratona)', 1103642.0),
   ('Perchero de pie ×1', 40000.0), ('Tacho de basura grande ×1', 56300.0)]),
 ('MOBILIARIO · escenario y técnica', [
   ('Mesa cocktail alta (escenario) ×1', 45000.0), ('Mesa alta para DJ ×1', 62300.0)]),
 ('GENERALES', [
   ('Dirección técnica de todo el evento + replanteo', 5149179.20),
   ('Guardia eléctrica × 2 días', 865800.0), ('Guardia técnica × 2 días', 905200.0)]),
]
PROVISIONAL = {1, 2, 3, 26, 27, 28}   # panelería + generales: el escenario que está cargado hoy

ws = hoja('INFRAESTRUCTURA', 'Presupuesto de infraestructura de La Rural, línea por línea',
          'Las 28 líneas del PDF del 17/08. Marcá SÍ o NO en la columna "¿Se queda?" y el subtotal se recalcula.')
COLS = [('#', 5), ('Bloque', 30), ('Ítem', 48), ('Precio (ARS)', 15), ('USD', 12),
        ('¿Se queda?', 12), ('Si se queda (ARS)', 17), ('Nota', 44)]
NC = len(COLS)
encabezados(ws, 4, COLS)
dv = DataValidation(type='list', formula1='"SÍ,NO"', allow_blank=True)
ws.add_data_validation(dv)
NOTAS = {
 3: 'Depósito para las 2.500 barras de la dinámica.',
 15: 'Reemplazadas por las 5.500 sillas de FDL Eventos.',
 21: 'El mobiliario "más fino" del camarín: candidato a contratárselo a La Rural.',
 26: 'Ojo: el replanteo de sillas va aparte, con ingeniero propio.',
}
r, n, filas_infra = 5, 0, []
for bloque, items in INFRA:
    r = banda_bloque(ws, r, bloque, NC)
    for desc, precio in items:
        n += 1
        ws.cell(r, 1, n).font = SUB
        ws.cell(r, 2, bloque.split('·')[-1].strip()).font = SUB
        ws.cell(r, 3, desc).font = AZUL
        c = ws.cell(r, 4, precio); c.font = AZUL; c.number_format = ARS
        c = ws.cell(r, 5, f'=D{r}/TABLERO!$C$4'); c.font = NEGRO; c.number_format = USD
        c = ws.cell(r, 6, 'SÍ' if n in PROVISIONAL else 'NO'); c.font = AZUL; c.fill = FIN
        dv.add(c)
        c = ws.cell(r, 7, f'=IF(F{r}="SÍ",D{r},0)'); c.font = NEGRO; c.number_format = ARS
        ws.cell(r, 8, NOTAS.get(n, '')).font = SUB
        for j in range(1, NC + 1): ws.cell(r, j).border = BOX
        filas_infra.append(r)
        r += 1
ini_i, fin_i = filas_infra[0], filas_infra[-1]
r += 1
for etiqueta, formula, relleno in [
    ('TOTAL del presupuesto de La Rural', f'=SUM(D{ini_i}:D{fin_i})', FGRIS),
    ('Lo que se queda (marcado SÍ)',      f'=SUM(G{ini_i}:G{fin_i})', FVERD),
    ('Lo que se saca (marcado NO)',       f'=D{r}-D{r+1}',            FAMAR)]:
    ws.cell(r, 3, etiqueta).font = BOLD
    c = ws.cell(r, 4, formula); c.font = BOLD; c.number_format = ARS; c.fill = relleno; c.border = BOX
    c = ws.cell(r, 5, f'=D{r}/TABLERO!$C$4'); c.font = BOLD; c.number_format = USD; c.fill = relleno; c.border = BOX
    r += 1
ws.cell(r + 1, 3, 'Lo marcado hoy es el escenario provisional cargado en COSTOS UNIFICADOS: panelería + dirección técnica + guardias. Cambiá los SÍ/NO y llevá el nuevo total a esa hoja.').font = SUB

del wb['Sheet']
wb.save('ARGENTINA_CMC2026_Costo_Total.xlsx')
print('OK · fila total en COSTOS UNIFICADOS:', fila_total)
